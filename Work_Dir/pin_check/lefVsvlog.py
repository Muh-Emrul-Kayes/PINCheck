import os
import json
import sys
import re
from log import Logging
from utils import regexExtraction, write, jsonWrite,check_file, read,jsonRead, makeDirs,check_file
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment,PatternFill,Border,Side,Font 
from openpyxl.worksheet.table import Table, TableStyleInfo

# LEF Vs VLOG Compare 
def compareLefVsVlog(wb,ws,lefFileKeys, vlogFileKeys, lefJson, vlogJson, IP_TYPE, ExcelKeys):
    """
    This function will generate the lef vs vlog comparison report as spreadsheet.

    Usage::
        >>> import lefVsvlog
        >>> lefVsvlog.compareLefVsVlog("WorkBook","WorkSheet","lef_file_list","vlog_file_list",
                                        "lef_json_data", "vlog_json_data", "IP_TYPE", "Excel_heading")

    :param: wb: workbook 
    :param: ws: worksheet
    :param: lefFileKeys: lef file list 
    :param: vlogFileKeys: vlog file list
    :param: lefJson: lef data as json dictionary format
    :param: vlogJson: vlog data as json dictionary format
    :param: IP_TYPE: ip type. Ex. memory, io, logic etc.
    :param: ExcelKeys: spreadsheet heading as list 
    :returns: wb: workbook
    :returns: ws: worksheet
    """  
    Logging.message("INFO", "GENERATING LEF Vs VERILOG COMPARISON REPORT")
    for lefFile in lefFileKeys:
        for vlogFile in vlogFileKeys:
            lefCellKeys = [*lefJson[lefFile].keys()]
            vlogCellKeys = [*vlogJson[vlogFile].keys()]
            MergeCellKeys = sorted(list(set(lefCellKeys)|set(vlogCellKeys)))
            for CellKey in MergeCellKeys:
                if CellKey in lefCellKeys and CellKey in vlogCellKeys:
                    lefPortKeys = [*lefJson[lefFile][CellKey].keys()]
                    vlogPortKeys = [*vlogJson[vlogFile][CellKey].keys()]
                    MergePortKeys = sorted(list(set(lefPortKeys)|set(vlogPortKeys)))
                    # IF PORT NOT FOUND IN LEF AND VLOG > COMMENT
                    if len(MergePortKeys) == 0:
                        #
                        FoundExistKey = False
                        for rowNum, rowWrite in enumerate(ws.iter_rows(values_only=True), start=1):
                            if rowWrite[0] == CellKey and rowWrite[1] == "NA":
                                # LEF DIR
                                ws.cell(row=rowNum,column=3,value="NA").font = Font(color='000000FF')
                                # vlog DIR
                                ws.cell(row=rowNum,column=5,value="NA").font = Font(color='000000FF')
                                # ERROR COMMENT
                                if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                    ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                else :
                                    ExistComment = ""
                                Commentlef = f"ERROR : For Cell '{CellKey}' Pin not defined in lef"
                                commentvlog = f"ERROR : For Cell '{CellKey}' Pin not defined in vlog"
                                ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {vlogFile}\n{Commentlef}\n{commentvlog}")).font = Font(color='FF000000')
                                #                                                
                                # STATUS
                                ws.cell(row=rowNum,column=7,value="NA").font = Font(color='000000FF')
                                FoundExistKey = True
                                break                        
                        if not FoundExistKey:
                            FoundExistCell = False
                            for rowNum, rowWrite in enumerate(ws.iter_rows(values_only=True), start=1):
                                if rowWrite[0] == CellKey:
                                    ws.insert_rows(rowNum)
                                    ws.cell(row=rowNum,column=1,value=str(CellKey)).font = Font(color='FF000000')
                                    ws.cell(row=rowNum,column=2,value="NA").font = Font(color='000000FF')
                                    # LEF DIR
                                    ws.cell(row=rowNum,column=3,value="NA").font = Font(color='000000FF')
                                    # vlog DIR
                                    ws.cell(row=rowNum,column=5,value="NA").font = Font(color='000000FF')
                                    # ERROR COMMENT
                                    if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                        ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                    else :
                                        ExistComment = ""
                                    Commentlef = f"ERROR : For Cell '{CellKey}' Pin not defined in lef"
                                    commentvlog = f"ERROR : For Cell '{CellKey}' Pin not defined in vlog"
                                    ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {vlogFile}\n{Commentlef}\n{commentvlog}")).font = Font(color='FF000000')
                                    #                                                 
                                    ws.cell(row=rowNum,column=7,value="NA").font = Font(color='000000FF')
                                    FoundExistCell = True
                                    break 
                            if not FoundExistCell:
                                rowNum = len(ws['A'])+1
                                ws.cell(row=rowNum,column=1,value=str(CellKey)).font = Font(color='FF000000')
                                ws.cell(row=rowNum,column=2,value="NA").font = Font(color='000000FF')
                                # LEF DIR
                                ws.cell(row=rowNum,column=3,value="NA").font = Font(color='000000FF')
                                # vlog DIR
                                ws.cell(row=rowNum,column=5,value="NA").font = Font(color='000000FF')
                                # ERROR COMMENT
                                if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                    ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                else :
                                    ExistComment = ""
                                Commentlef = f"ERROR : For Cell '{CellKey}' Pin not defined in lef"
                                commentvlog = f"ERROR : For Cell '{CellKey}' Pin not defined in vlog"
                                ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {vlogFile}\n{Commentlef}\n{commentvlog}")).font = Font(color='FF000000')
                                #                                                 
                                ws.cell(row=rowNum,column=7,value="NA").font = Font(color='000000FF')
                        #
                    # Compare PIN Between LEF and VLOG Cell
                    for PortKey in MergePortKeys:
                        if PortKey in lefPortKeys and PortKey in vlogPortKeys:
                            lefPortDir = lefJson[lefFile][CellKey][PortKey]["PortDirection"].upper()
                            vlogPortDir = vlogJson[vlogFile][CellKey][PortKey]["PortDirection"].upper()
                            if lefPortDir == vlogPortDir and lefPortDir.strip() and vlogPortDir.strip():
                                #
                                FoundExistKey = False
                                for rowNum, rowWrite in enumerate(ws.iter_rows(values_only=True), start=1):
                                    if rowWrite[0] == CellKey and rowWrite[1] == PortKey:
                                        # LEF DIR
                                        if (rowWrite[2] == None or rowWrite[2] == "NA") and rowWrite[2] != "ERROR":
                                            if lefPortDir.strip():
                                                ws.cell(row=rowNum,column=3,value=str(lefPortDir.upper())).font = Font(color='FF000000')
                                            else:
                                                ws.cell(row=rowNum,column=3,value="ERROR").font = Font(color='00FF0000')
                                                # ERROR COMMENT
                                                if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                    ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                                else :
                                                    ExistComment = ""
                                                ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in lef")).font = Font(color='FF000000')
                                                #
                                        else : 
                                            if rowWrite[2] == lefPortDir.upper():
                                                pass
                                            else :
                                                ws.cell(row=rowNum,column=3,value="ERROR").font = Font(color='00FF0000')
                                                # ERROR COMMENT
                                                if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                    ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                                else :
                                                    ExistComment = ""
                                                if lefPortDir.strip():
                                                    ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction mismatched within lef")).font = Font(color='FF000000')
                                                else:
                                                    ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in lef")).font = Font(color='FF000000')
                                                #
                                        # VLOG DIR
                                        if (rowWrite[4] == None or rowWrite[4] == "NA") and rowWrite[4] != "ERROR":
                                            if vlogPortDir.strip():
                                                ws.cell(row=rowNum,column=5,value=str(vlogPortDir.upper())).font = Font(color='FF000000')
                                            else:
                                                ws.cell(row=rowNum,column=5,value="ERROR").font = Font(color='00FF0000')
                                                # ERROR COMMENT
                                                if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                    ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                                else :
                                                    ExistComment = ""
                                                ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in vlog")).font = Font(color='FF000000')
                                                #                                                
                                        else : 
                                            if rowWrite[4] == vlogPortDir.upper():
                                                pass
                                            else :
                                                ws.cell(row=rowNum,column=5,value="ERROR").font = Font(color='00FF0000') 
                                                # ERROR COMMENT
                                                if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                    ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                                else :
                                                    ExistComment = ""
                                                if vlogPortDir.strip():
                                                    ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction mismatched within vlog")).font = Font(color='FF000000')
                                                else:
                                                    ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in vlog")).font = Font(color='FF000000')
                                                #                                                
                                        # STATUS
                                        if rowWrite[6] == None or rowWrite[6] == "NA":
                                            ws.cell(row=rowNum,column=7,value="PASS").font = Font(color='00008000')
                                        FoundExistKey = True
                                        break                        
                                if not FoundExistKey:
                                    FoundExistCell = False
                                    for rowNum, rowWrite in enumerate(ws.iter_rows(values_only=True), start=1):
                                        if rowWrite[0] == CellKey:
                                            ws.insert_rows(rowNum)
                                            ws.cell(row=rowNum,column=1,value=str(CellKey)).font = Font(color='FF000000')
                                            ws.cell(row=rowNum,column=2,value=str(PortKey)).font = Font(color='FF000000')
                                            if lefPortDir.strip():
                                                ws.cell(row=rowNum,column=3,value=str(lefPortDir.upper())).font = Font(color='FF000000')
                                            else:
                                                ws.cell(row=rowNum,column=3,value="ERROR").font = Font(color='00FF0000')
                                                # ERROR COMMENT
                                                if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                    ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                                else :
                                                    ExistComment = ""
                                                ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in lef")).font = Font(color='FF000000')
                                                #                                                 
                                            if vlogPortDir.strip():
                                                ws.cell(row=rowNum,column=5,value=str(vlogPortDir.upper())).font = Font(color='FF000000')
                                            else:
                                                ws.cell(row=rowNum,column=5,value="ERROR").font = Font(color='00FF0000')
                                                # ERROR COMMENT
                                                if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                    ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                                else :
                                                    ExistComment = ""
                                                ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in vlog")).font = Font(color='FF000000')
                                                #                                                 
                                            ws.cell(row=rowNum,column=7,value="PASS").font = Font(color='00008000')
                                            FoundExistCell = True
                                            break 
                                    if not FoundExistCell:
                                        rowNum = len(ws['A'])+1
                                        ws.cell(row=rowNum,column=1,value=str(CellKey)).font = Font(color='FF000000')
                                        ws.cell(row=rowNum,column=2,value=str(PortKey)).font = Font(color='FF000000')
                                        if lefPortDir.strip():
                                            ws.cell(row=rowNum,column=3,value=str(lefPortDir.upper())).font = Font(color='FF000000')
                                        else:
                                            ws.cell(row=rowNum,column=3,value="ERROR").font = Font(color='00FF0000')
                                            # ERROR COMMENT
                                            if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                            else :
                                                ExistComment = ""
                                            ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in lef")).font = Font(color='FF000000')
                                            #                                             
                                        if vlogPortDir.strip():
                                            ws.cell(row=rowNum,column=5,value=str(vlogPortDir.upper())).font = Font(color='FF000000')
                                        else:
                                            ws.cell(row=rowNum,column=5,value="ERROR").font = Font(color='00FF0000')
                                            # ERROR COMMENT
                                            if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                            else :
                                                ExistComment = ""
                                            ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in vlog")).font = Font(color='FF000000')
                                            #                                             
                                        ws.cell(row=rowNum,column=7,value="PASS").font = Font(color='00008000')
                                #                                                      
                            else:
                                #
                                FoundExistKey = False
                                for rowNum, rowWrite in enumerate(ws.iter_rows(values_only=True), start=1):
                                    if rowWrite[0] == CellKey and rowWrite[1] == PortKey:
                                        # LEF DIR
                                        if (rowWrite[2] == None or rowWrite[2] == "NA") and rowWrite[2] != "ERROR":
                                            if lefPortDir.strip():
                                                ws.cell(row=rowNum,column=3,value=str(lefPortDir.upper())).font = Font(color='FF000000')
                                            else:
                                                ws.cell(row=rowNum,column=3,value="ERROR").font = Font(color='00FF0000')
                                                # ERROR COMMENT
                                                if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                    ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                                else :
                                                    ExistComment = ""
                                                ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in lef")).font = Font(color='FF000000')
                                                #                                                 
                                        else : 
                                            if rowWrite[2] == lefPortDir.upper():
                                                pass
                                            else :
                                                ws.cell(row=rowNum,column=3,value="ERROR").font = Font(color='00FF0000')
                                                # ERROR COMMENT
                                                if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                    ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                                else :
                                                    ExistComment = ""
                                                if lefPortDir.strip():
                                                    ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction mismatched within lef")).font = Font(color='FF000000')
                                                else:
                                                    ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in lef")).font = Font(color='FF000000')
                                                #                                                
                                        # VLOG DIR
                                        if (rowWrite[4] == None or rowWrite[4] == "NA") and rowWrite[4] != "ERROR":
                                            if vlogPortDir.strip():
                                                ws.cell(row=rowNum,column=5,value=str(vlogPortDir.upper())).font = Font(color='FF000000')
                                            else:
                                                ws.cell(row=rowNum,column=5,value="ERROR").font = Font(color='00FF0000')
                                                # ERROR COMMENT
                                                if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                    ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                                else :
                                                    ExistComment = ""
                                                ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in vlog")).font = Font(color='FF000000')
                                                #                                                 
                                        else : 
                                            if rowWrite[4] == vlogPortDir.upper():
                                                pass
                                            else :
                                                ws.cell(row=rowNum,column=5,value="ERROR").font = Font(color='00FF0000')
                                                # ERROR COMMENT
                                                if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                    ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                                else :
                                                    ExistComment = ""
                                                if vlogPortDir.strip():
                                                    ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction mismatched within vlog")).font = Font(color='FF000000')
                                                else:
                                                    ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in vlog")).font = Font(color='FF000000')
                                                #                                                 
                                        # STATUS
                                        if rowWrite[6] == None or rowWrite[6] == "NA" or rowWrite[6] == "PASS":
                                            ws.cell(row=rowNum,column=7,value="ERROR").font = Font(color='00FF0000')
                                        # COMMENT
                                        if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                            ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                        else :
                                            ExistComment = ""
                                        ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction mismatched between lef and vlog")).font = Font(color='FF000000')                                                                                        
                                        FoundExistKey = True
                                        break                        
                                if not FoundExistKey:
                                    FoundExistCell = False
                                    for rowNum, rowWrite in enumerate(ws.iter_rows(values_only=True), start=1):
                                        if rowWrite[0] == CellKey:
                                            ws.insert_rows(rowNum)
                                            ws.cell(row=rowNum,column=1,value=str(CellKey)).font = Font(color='FF000000')
                                            ws.cell(row=rowNum,column=2,value=str(PortKey)).font = Font(color='FF000000')
                                            if lefPortDir.strip():
                                                ws.cell(row=rowNum,column=3,value=str(lefPortDir.upper())).font = Font(color='FF000000')
                                            else:
                                                ws.cell(row=rowNum,column=3,value="ERROR").font = Font(color='00FF0000')
                                                # ERROR COMMENT
                                                if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                    ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                                else :
                                                    ExistComment = ""
                                                ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in lef")).font = Font(color='FF000000')
                                                #                                                 
                                            if vlogPortDir.strip():
                                                ws.cell(row=rowNum,column=5,value=str(vlogPortDir.upper())).font = Font(color='FF000000')
                                            else:
                                                ws.cell(row=rowNum,column=5,value="ERROR").font = Font(color='00FF0000')
                                                # ERROR COMMENT
                                                if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                    ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                                else :
                                                    ExistComment = ""
                                                ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in vlog")).font = Font(color='FF000000')
                                                #                                                 
                                            ws.cell(row=rowNum,column=7,value="ERROR").font = Font(color='00FF0000')
                                            # COMMENT
                                            if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                            else :
                                                ExistComment = ""
                                            ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction mismatched between lef and vlog")).font = Font(color='FF000000')                                            
                                            FoundExistCell = True
                                            break 
                                    if not FoundExistCell:
                                        rowNum = len(ws['A'])+1
                                        ws.cell(row=rowNum,column=1,value=str(CellKey)).font = Font(color='FF000000')
                                        ws.cell(row=rowNum,column=2,value=str(PortKey)).font = Font(color='FF000000')
                                        if lefPortDir.strip():
                                            ws.cell(row=rowNum,column=3,value=str(lefPortDir.upper())).font = Font(color='FF000000')
                                        else:
                                            ws.cell(row=rowNum,column=3,value="ERROR").font = Font(color='00FF0000')
                                            # ERROR COMMENT
                                            if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                            else :
                                                ExistComment = ""
                                            ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in lef")).font = Font(color='FF000000')
                                            #                                             
                                        if vlogPortDir.strip():
                                            ws.cell(row=rowNum,column=5,value=str(vlogPortDir.upper())).font = Font(color='FF000000')
                                        else:
                                            ws.cell(row=rowNum,column=5,value="ERROR").font = Font(color='00FF0000')
                                            # ERROR COMMENT
                                            if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                            else :
                                                ExistComment = ""
                                            ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in vlog")).font = Font(color='FF000000')
                                            #                                             
                                        ws.cell(row=rowNum,column=7,value="ERROR").font = Font(color='00FF0000')
                                        # COMMENT
                                        if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                            ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                        else :
                                            ExistComment = ""
                                        ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction mismatched between lef and vlog")).font = Font(color='FF000000')                                         
                                #                                
                        # FOR PORT NOT FOUND IN CELL BETWEEN LEF AND VLOG
                        else:
                            if PortKey in lefPortKeys:
                                lefPortDir = lefJson[lefFile][CellKey][PortKey]["PortDirection"].upper()  
                                #
                                FoundExistKey = False
                                for rowNum, rowWrite in enumerate(ws.iter_rows(values_only=True), start=1):
                                    if rowWrite[0] == CellKey and rowWrite[1] == PortKey:
                                        # LEF DIR
                                        if (rowWrite[2] == None or rowWrite[2] == "NA") and rowWrite[2] != "ERROR":
                                            if lefPortDir.strip():
                                                ws.cell(row=rowNum,column=3,value=str(lefPortDir.upper())).font = Font(color='FF000000')
                                            else:
                                                ws.cell(row=rowNum,column=3,value="ERROR").font = Font(color='00FF0000')
                                                # ERROR COMMENT
                                                if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                    ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                                else :
                                                    ExistComment = ""
                                                ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in lef")).font = Font(color='FF000000')
                                                #                                                 
                                        else : 
                                            if rowWrite[2] == lefPortDir.upper():
                                                pass
                                            else :
                                                ws.cell(row=rowNum,column=3,value="ERROR").font = Font(color='00FF0000')
                                                # ERROR COMMENT
                                                if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                    ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                                else :
                                                    ExistComment = ""
                                                if lefPortDir.strip():
                                                    ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction mismatched within lef")).font = Font(color='FF000000')
                                                else:
                                                    ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in lef")).font = Font(color='FF000000')
                                                #                                                
                                        # VLOG DIR
                                        if rowWrite[4] == None:
                                            ws.cell(row=rowNum,column=5,value="NA").font = Font(color='000000FF') 
                                        # STATUS
                                        # if rowWrite[6] != "PASS":
                                        #     ws.cell(row=rowNum,column=7,value="FAIL").font = Font(color='00FF0000')
                                        ws.cell(row=rowNum,column=7,value="FAIL").font = Font(color='00FF0000')
                                                                                   
                                        # COMMENT
                                        if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                            ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                        else :
                                            ExistComment = ""
                                        ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' defined in lef but not in vlog")).font = Font(color='FF000000')
                                        FoundExistKey = True
                                        break                        
                                if not FoundExistKey:
                                    FoundExistCell = False
                                    for rowNum, rowWrite in enumerate(ws.iter_rows(values_only=True), start=1):
                                        if rowWrite[0] == CellKey:
                                            ws.insert_rows(rowNum)
                                            ws.cell(row=rowNum,column=1,value=str(CellKey)).font = Font(color='FF000000')
                                            ws.cell(row=rowNum,column=2,value=str(PortKey)).font = Font(color='FF000000')
                                            if lefPortDir.strip():
                                                ws.cell(row=rowNum,column=3,value=str(lefPortDir.upper())).font = Font(color='FF000000')
                                            else:
                                                ws.cell(row=rowNum,column=3,value="ERROR").font = Font(color='00FF0000')
                                                # ERROR COMMENT
                                                if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                    ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                                else :
                                                    ExistComment = ""
                                                ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in lef")).font = Font(color='FF000000')
                                                #                                                 
                                            ws.cell(row=rowNum,column=5,value="NA").font = Font(color='000000FF')
                                            ws.cell(row=rowNum,column=7,value="FAIL").font = Font(color='00FF0000')
                                            
                                            # COMMENT
                                            if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                            else :
                                                ExistComment = ""
                                            ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' defined in lef but not in vlog")).font = Font(color='FF000000')                                            
                                            FoundExistCell = True
                                            break 
                                    if not FoundExistCell:
                                        rowNum = len(ws['A'])+1
                                        ws.cell(row=rowNum,column=1,value=str(CellKey)).font = Font(color='FF000000')
                                        ws.cell(row=rowNum,column=2,value=str(PortKey)).font = Font(color='FF000000')
                                        if lefPortDir.strip():
                                            ws.cell(row=rowNum,column=3,value=str(lefPortDir.upper())).font = Font(color='FF000000')
                                        else:
                                            ws.cell(row=rowNum,column=3,value="ERROR").font = Font(color='00FF0000')
                                            # ERROR COMMENT
                                            if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                            else :
                                                ExistComment = ""
                                            ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in lef")).font = Font(color='FF000000')
                                            #                                             
                                        ws.cell(row=rowNum,column=5,value="NA").font = Font(color='000000FF')
                                        ws.cell(row=rowNum,column=7,value="FAIL").font = Font(color='00FF0000')
                                        
                                        # COMMENT
                                        if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                            ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                        else :
                                            ExistComment = ""
                                        ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' defined in lef but not in vlog")).font = Font(color='FF000000')                                            
                                #                                   
                            else:
                                vlogPortDir = vlogJson[vlogFile][CellKey][PortKey]["PortDirection"].upper()   
                                #
                                FoundExistKey = False
                                for rowNum, rowWrite in enumerate(ws.iter_rows(values_only=True), start=1):
                                    if rowWrite[0] == CellKey and rowWrite[1] == PortKey:
                                        # LEF DIR
                                        if rowWrite[2] == None:
                                            ws.cell(row=rowNum,column=3,value="NA").font = Font(color='000000FF')
                                        # VLOG DIR 
                                        if (rowWrite[4] == None or rowWrite[4] == "NA") and rowWrite[4] != "ERROR":
                                            if vlogPortDir.strip():
                                                ws.cell(row=rowNum,column=5,value=str(vlogPortDir.upper())).font = Font(color='FF000000')
                                            else:
                                                ws.cell(row=rowNum,column=5,value="ERROR").font = Font(color='00FF0000')
                                                # ERROR COMMENT
                                                if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                    ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                                else :
                                                    ExistComment = ""
                                                ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in vlog")).font = Font(color='FF000000')
                                                #                                                 
                                        else : 
                                            if rowWrite[4] == vlogPortDir.upper():
                                                pass
                                            else :
                                                ws.cell(row=rowNum,column=5,value="ERROR").font = Font(color='00FF0000')         
                                                # ERROR COMMENT
                                                if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                    ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                                else :
                                                    ExistComment = ""
                                                if vlogPortDir.strip():
                                                    ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction mismatched within vlog")).font = Font(color='FF000000')
                                                else:
                                                    ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in vlog")).font = Font(color='FF000000')
                                                #                                                                                   
                                        # STATUS
                                        # if rowWrite[6] != "PASS":
                                        #     ws.cell(row=rowNum,column=7,value="FAIL").font = Font(color='00FF0000')
                                        ws.cell(row=rowNum,column=7,value="FAIL").font = Font(color='00FF0000')
                                        
                                        # COMMENT
                                        if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                            ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                        else :
                                            ExistComment = ""
                                        ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' defined in vlog but not in lef")).font = Font(color='FF000000')
                                        FoundExistKey = True
                                        break                        
                                if not FoundExistKey:
                                    FoundExistCell = False
                                    for rowNum, rowWrite in enumerate(ws.iter_rows(values_only=True), start=1):
                                        if rowWrite[0] == CellKey:
                                            ws.insert_rows(rowNum)
                                            ws.cell(row=rowNum,column=1,value=str(CellKey)).font = Font(color='FF000000')
                                            ws.cell(row=rowNum,column=2,value=str(PortKey)).font = Font(color='FF000000')
                                            ws.cell(row=rowNum,column=3,value="NA").font = Font(color='000000FF')
                                            if vlogPortDir.strip():
                                                ws.cell(row=rowNum,column=5,value=str(vlogPortDir.upper())).font = Font(color='FF000000')
                                            else:
                                                ws.cell(row=rowNum,column=5,value="ERROR").font = Font(color='00FF0000')
                                                # ERROR COMMENT
                                                if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                    ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                                else :
                                                    ExistComment = ""
                                                ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in vlog")).font = Font(color='FF000000')
                                                #                                                 
                                            ws.cell(row=rowNum,column=7,value="FAIL").font = Font(color='00FF0000')
                                            
                                            # COMMENT
                                            if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                            else :
                                                ExistComment = ""
                                            ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' defined in vlog but not in lef")).font = Font(color='FF000000')                                            
                                            FoundExistCell = True
                                            break 
                                    if not FoundExistCell:
                                        rowNum = len(ws['A'])+1
                                        ws.cell(row=rowNum,column=1,value=str(CellKey)).font = Font(color='FF000000')
                                        ws.cell(row=rowNum,column=2,value=str(PortKey)).font = Font(color='FF000000')
                                        ws.cell(row=rowNum,column=3,value="NA").font = Font(color='000000FF')
                                        if vlogPortDir.strip():
                                            ws.cell(row=rowNum,column=5,value=str(vlogPortDir.upper())).font = Font(color='FF000000')
                                        else:
                                            ws.cell(row=rowNum,column=5,value="ERROR").font = Font(color='00FF0000')
                                            # ERROR COMMENT
                                            if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                            else :
                                                ExistComment = ""
                                            ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in vlog")).font = Font(color='FF000000')
                                            #                                             
                                        ws.cell(row=rowNum,column=7,value="FAIL").font = Font(color='00FF0000')
                                        
                                        # COMMENT
                                        if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                            ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                        else :
                                            ExistComment = ""
                                        ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' defined in vlog but not in lef")).font = Font(color='FF000000')                                            
                                #                                       
                # FOR CELL NOT FOUND IN ONE FILE LEF OR VLOG        
                else:
                    if CellKey in lefCellKeys:
                        lefPortKeys = [*lefJson[lefFile][CellKey].keys()]
                        # IF PORT NOT FOUND IN LEF > ERROR COMMENT
                        if len(lefPortKeys) == 0:
                            #
                            FoundExistKey = False
                            for rowNum, rowWrite in enumerate(ws.iter_rows(values_only=True), start=1):
                                if rowWrite[0] == CellKey and rowWrite[1] == "NA":                                  
                                    # lef DIR
                                    ws.cell(row=rowNum,column=3,value="NA").font = Font(color='000000FF')
                                    # vlog DIR
                                    ws.cell(row=rowNum,column=5,value="NA").font = Font(color='000000FF')
                                    # ERROR COMMENT
                                    if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                        ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                    else :
                                        ExistComment = ""
                                    Commentlef = f"ERROR : For Cell '{CellKey}' Pin not defined in lef"
                                    commentvlog = f"ERROR : Cell '{CellKey}' not defined in vlog"
                                    ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {vlogFile}\n{Commentlef}\n{commentvlog}")).font = Font(color='FF000000')
                                    #                                                
                                    # STATUS
                                    ws.cell(row=rowNum,column=7,value="FAIL").font = Font(color='00FF0000')
                                    FoundExistKey = True
                                    break                        
                            if not FoundExistKey:
                                FoundExistCell = False
                                for rowNum, rowWrite in enumerate(ws.iter_rows(values_only=True), start=1):
                                    if rowWrite[0] == CellKey:
                                        ws.insert_rows(rowNum)
                                        ws.cell(row=rowNum,column=1,value=str(CellKey)).font = Font(color='FF000000')
                                        ws.cell(row=rowNum,column=2,value="NA").font = Font(color='000000FF')
                                        # lef DIR
                                        ws.cell(row=rowNum,column=3,value="NA").font = Font(color='000000FF')
                                        # vlog DIR
                                        ws.cell(row=rowNum,column=5,value="NA").font = Font(color='000000FF')
                                        # ERROR COMMENT
                                        if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                            ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                        else :
                                            ExistComment = ""
                                        Commentlef = f"ERROR : For Cell '{CellKey}' Pin not defined in lef"
                                        commentvlog = f"ERROR : Cell '{CellKey}' not defined in vlog"
                                        ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {vlogFile}\n{Commentlef}\n{commentvlog}")).font = Font(color='FF000000')
                                        #                                                
                                        # STATUS
                                        ws.cell(row=rowNum,column=7,value="FAIL").font = Font(color='00FF0000')
                                        FoundExistCell = True
                                        break 
                                if not FoundExistCell:
                                    rowNum = len(ws['A'])+1
                                    ws.cell(row=rowNum,column=1,value=str(CellKey)).font = Font(color='FF000000')
                                    ws.cell(row=rowNum,column=2,value="NA").font = Font(color='000000FF')
                                    # lef DIR
                                    ws.cell(row=rowNum,column=3,value="NA").font = Font(color='000000FF')
                                    # vlog DIR
                                    ws.cell(row=rowNum,column=5,value="NA").font = Font(color='000000FF')
                                    # ERROR COMMENT
                                    if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                        ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                    else :
                                        ExistComment = ""
                                    Commentlef = f"ERROR : For Cell '{CellKey}' Pin not defined in lef"
                                    commentvlog = f"ERROR : Cell '{CellKey}' not defined in vlog"
                                    ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {vlogFile}\n{Commentlef}\n{commentvlog}")).font = Font(color='FF000000')
                                    #                                                
                                    # STATUS
                                    ws.cell(row=rowNum,column=7,value="FAIL").font = Font(color='00FF0000')
                            #                       
                        for PortKey in lefPortKeys:
                            lefPortDir = lefJson[lefFile][CellKey][PortKey]["PortDirection"].upper()   
                            #
                            FoundExistKey = False
                            for rowNum, rowWrite in enumerate(ws.iter_rows(values_only=True), start=1):
                                if rowWrite[0] == CellKey and rowWrite[1] == PortKey:
                                    # LEF DIR
                                    if (rowWrite[2] == None or rowWrite[2] == "NA") and rowWrite[2] != "ERROR":
                                        if lefPortDir.strip():
                                            ws.cell(row=rowNum,column=3,value=str(lefPortDir.upper())).font = Font(color='FF000000')
                                        else:
                                            ws.cell(row=rowNum,column=3,value="ERROR").font = Font(color='00FF0000')
                                            # ERROR COMMENT
                                            if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                            else :
                                                ExistComment = ""
                                            ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in lef")).font = Font(color='FF000000')
                                            #                                             
                                    else : 
                                        if rowWrite[2] == lefPortDir.upper():
                                            pass
                                        else :
                                            ws.cell(row=rowNum,column=3,value="ERROR").font = Font(color='00FF0000')  
                                            # ERROR COMMENT
                                            if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                            else :
                                                ExistComment = ""
                                            if lefPortDir.strip():
                                                ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction mismatched within lef")).font = Font(color='FF000000')
                                            else:
                                                ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in lef")).font = Font(color='FF000000')
                                            #                                                                               
                                    # VLOG DIR 
                                    if rowWrite[4] == None:
                                        ws.cell(row=rowNum,column=5,value="NA").font = Font(color='000000FF')                                             
                                    # STATUS
                                    # if rowWrite[6] != "PASS":
                                    #     ws.cell(row=rowNum,column=7,value="FAIL").font = Font(color='00FF0000')
                                    ws.cell(row=rowNum,column=7,value="FAIL").font = Font(color='00FF0000')  
                                    
                                    # COMMENT
                                    if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                        ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                    else :
                                        ExistComment = ""
                                    ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {vlogFile}\nERROR : Cell '{CellKey}' defined in lef but not in vlog")).font = Font(color='FF000000')
                                    FoundExistKey = True
                                    break                        
                            if not FoundExistKey:
                                FoundExistCell = False
                                for rowNum, rowWrite in enumerate(ws.iter_rows(values_only=True), start=1):
                                    if rowWrite[0] == CellKey:
                                        ws.insert_rows(rowNum)
                                        ws.cell(row=rowNum,column=1,value=str(CellKey)).font = Font(color='FF000000')
                                        ws.cell(row=rowNum,column=2,value=str(PortKey)).font = Font(color='FF000000')
                                        if lefPortDir.strip():
                                            ws.cell(row=rowNum,column=3,value=str(lefPortDir.upper())).font = Font(color='FF000000')
                                        else:
                                            ws.cell(row=rowNum,column=3,value="ERROR").font = Font(color='00FF0000')
                                            # ERROR COMMENT
                                            if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                            else :
                                                ExistComment = ""
                                            ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in lef")).font = Font(color='FF000000')
                                            #                                            
                                        ws.cell(row=rowNum,column=5,value="NA").font = Font(color='000000FF')
                                        ws.cell(row=rowNum,column=7,value="FAIL").font = Font(color='00FF0000')
                                        
                                        # COMMENT
                                        if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                            ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                        else :
                                            ExistComment = ""
                                        ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {vlogFile}\nERROR : Cell '{CellKey}' defined in lef but not in vlog")).font = Font(color='FF000000')                                            
                                        FoundExistCell = True
                                        break 
                                if not FoundExistCell:
                                    rowNum = len(ws['A'])+1
                                    ws.cell(row=rowNum,column=1,value=str(CellKey)).font = Font(color='FF000000')
                                    ws.cell(row=rowNum,column=2,value=str(PortKey)).font = Font(color='FF000000')
                                    if lefPortDir.strip():
                                        ws.cell(row=rowNum,column=3,value=str(lefPortDir.upper())).font = Font(color='FF000000')
                                    else:
                                        ws.cell(row=rowNum,column=3,value="ERROR").font = Font(color='00FF0000')
                                        # ERROR COMMENT
                                        if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                            ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                        else :
                                            ExistComment = ""
                                        ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in lef")).font = Font(color='FF000000')
                                        #                                        
                                    ws.cell(row=rowNum,column=5,value="NA").font = Font(color='000000FF')
                                    ws.cell(row=rowNum,column=7,value="FAIL").font = Font(color='00FF0000')
                                    
                                    # COMMENT
                                    if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                        ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                    else :
                                        ExistComment = ""
                                    ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {vlogFile}\nERROR : Cell '{CellKey}' defined in lef but not in vlog")).font = Font(color='FF000000')                                            
                            #                        
                    else: 
                        vlogPortKeys = [*vlogJson[vlogFile][CellKey].keys()]
                        # IF PORT NOT FOUND IN VLOG > ERROR COMMENT
                        if len(vlogPortKeys) == 0:
                            #
                            FoundExistKey = False
                            for rowNum, rowWrite in enumerate(ws.iter_rows(values_only=True), start=1):
                                if rowWrite[0] == CellKey and rowWrite[1] == "NA":
                                    # lef DIR
                                    ws.cell(row=rowNum,column=3,value="NA").font = Font(color='000000FF')
                                    # vlog DIR
                                    ws.cell(row=rowNum,column=5,value="NA").font = Font(color='000000FF')
                                    # ERROR COMMENT
                                    if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                        ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                    else :
                                        ExistComment = ""
                                    Commentlef = f"ERROR : Cell '{CellKey}' not defined in lef"
                                    commentvlog = f"ERROR : For Cell '{CellKey}' Pin not defined in vlog"
                                    ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {vlogFile}\n{Commentlef}\n{commentvlog}")).font = Font(color='FF000000')
                                    #                                                
                                    # STATUS
                                    ws.cell(row=rowNum,column=7,value="FAIL").font = Font(color='00FF0000')
                                    FoundExistKey = True
                                    break                        
                            if not FoundExistKey:
                                FoundExistCell = False
                                for rowNum, rowWrite in enumerate(ws.iter_rows(values_only=True), start=1):
                                    if rowWrite[0] == CellKey:
                                        ws.insert_rows(rowNum)
                                        ws.cell(row=rowNum,column=1,value=str(CellKey)).font = Font(color='FF000000')
                                        ws.cell(row=rowNum,column=2,value="NA").font = Font(color='000000FF')
                                        # lef DIR
                                        ws.cell(row=rowNum,column=3,value="NA").font = Font(color='000000FF')
                                        # vlog DIR
                                        ws.cell(row=rowNum,column=5,value="NA").font = Font(color='000000FF')
                                        # ERROR COMMENT
                                        if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                            ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                        else :
                                            ExistComment = ""
                                        Commentlef = f"ERROR : Cell '{CellKey}' not defined in lef"
                                        commentvlog = f"ERROR : For Cell '{CellKey}' Pin not defined in vlog"
                                        ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {vlogFile}\n{Commentlef}\n{commentvlog}")).font = Font(color='FF000000')
                                        #                                                
                                        # STATUS
                                        ws.cell(row=rowNum,column=7,value="FAIL").font = Font(color='00FF0000')
                                        FoundExistCell = True
                                        break 
                                if not FoundExistCell:
                                    rowNum = len(ws['A'])+1
                                    ws.cell(row=rowNum,column=1,value=str(CellKey)).font = Font(color='FF000000')
                                    ws.cell(row=rowNum,column=2,value="NA").font = Font(color='000000FF')
                                    # lef DIR
                                    ws.cell(row=rowNum,column=3,value="NA").font = Font(color='000000FF')
                                    # vlog DIR
                                    ws.cell(row=rowNum,column=5,value="NA").font = Font(color='000000FF')
                                    # ERROR COMMENT
                                    if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                        ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                    else :
                                        ExistComment = ""
                                    Commentlef = f"ERROR : Cell '{CellKey}' not defined in lef"
                                    commentvlog = f"ERROR : For Cell '{CellKey}' Pin not defined in vlog"
                                    ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {vlogFile}\n{Commentlef}\n{commentvlog}")).font = Font(color='FF000000')
                                    #                                                
                                    # STATUS
                                    ws.cell(row=rowNum,column=7,value="FAIL").font = Font(color='00FF0000')
                            #                         
                        for PortKey in vlogPortKeys:
                            vlogPortDir = vlogJson[vlogFile][CellKey][PortKey]["PortDirection"].upper()   
                            #
                            FoundExistKey = False
                            for rowNum, rowWrite in enumerate(ws.iter_rows(values_only=True), start=1):
                                if rowWrite[0] == CellKey and rowWrite[1] == PortKey:
                                    # LEF DIR
                                    if rowWrite[2] == None:
                                        ws.cell(row=rowNum,column=3,value="NA").font = Font(color='000000FF')                                   
                                    # VLOG DIR 
                                    if (rowWrite[4] == None or rowWrite[4] == "NA") and rowWrite[4] != "ERROR":
                                        if vlogPortDir.strip():
                                            ws.cell(row=rowNum,column=5,value=str(vlogPortDir.upper())).font = Font(color='FF000000')
                                        else:
                                            ws.cell(row=rowNum,column=5,value="ERROR").font = Font(color='00FF0000')
                                            # ERROR COMMENT
                                            if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                            else :
                                                ExistComment = ""
                                            ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in vlog")).font = Font(color='FF000000')
                                            #                                            
                                    else : 
                                        if rowWrite[4] == vlogPortDir.upper():
                                            pass
                                        else :
                                            ws.cell(row=rowNum,column=5,value="ERROR").font = Font(color='00FF0000')
                                            # ERROR COMMENT
                                            if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                            else :
                                                ExistComment = ""
                                            if vlogPortDir.strip():
                                                ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction mismatched within vlog")).font = Font(color='FF000000')
                                            else:
                                                ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in vlog")).font = Font(color='FF000000')
                                            #
                                    # STATUS
                                    # if rowWrite[6] != "PASS":
                                    #     ws.cell(row=rowNum,column=7,value="FAIL").font = Font(color='00FF0000')
                                    ws.cell(row=rowNum,column=7,value="FAIL").font = Font(color='00FF0000')  
                                    
                                    # COMMENT
                                    if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                        ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                    else :
                                        ExistComment = ""
                                    ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {vlogFile}\nERROR : Cell '{CellKey}' defined in vlog but not in lef")).font = Font(color='FF000000')
                                    FoundExistKey = True
                                    break                        
                            if not FoundExistKey:
                                FoundExistCell = False
                                for rowNum, rowWrite in enumerate(ws.iter_rows(values_only=True), start=1):
                                    if rowWrite[0] == CellKey:
                                        ws.insert_rows(rowNum)
                                        ws.cell(row=rowNum,column=1,value=str(CellKey)).font = Font(color='FF000000')
                                        ws.cell(row=rowNum,column=2,value=str(PortKey)).font = Font(color='FF000000')
                                        ws.cell(row=rowNum,column=3,value="NA").font = Font(color='000000FF')
                                        if vlogPortDir.strip():
                                            ws.cell(row=rowNum,column=5,value=str(vlogPortDir.upper())).font = Font(color='FF000000')
                                        else:
                                            ws.cell(row=rowNum,column=5,value="ERROR").font = Font(color='00FF0000')
                                            # ERROR COMMENT
                                            if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                            else :
                                                ExistComment = ""
                                            ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in vlog")).font = Font(color='FF000000')
                                            #                                             
                                        ws.cell(row=rowNum,column=7,value="FAIL").font = Font(color='00FF0000')
                                        
                                        # COMMENT
                                        if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                            ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                        else :
                                            ExistComment = ""
                                        ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {vlogFile}\nERROR : Cell '{CellKey}' defined in vlog but not in lef")).font = Font(color='FF000000')                                            
                                        FoundExistCell = True
                                        break 
                                if not FoundExistCell:
                                    rowNum = len(ws['A'])+1
                                    ws.cell(row=rowNum,column=1,value=str(CellKey)).font = Font(color='FF000000')
                                    ws.cell(row=rowNum,column=2,value=str(PortKey)).font = Font(color='FF000000')
                                    ws.cell(row=rowNum,column=3,value="NA").font = Font(color='000000FF')
                                    if vlogPortDir.strip():
                                        ws.cell(row=rowNum,column=5,value=str(vlogPortDir.upper())).font = Font(color='FF000000')
                                    else:
                                        ws.cell(row=rowNum,column=5,value="ERROR").font = Font(color='00FF0000')
                                        # ERROR COMMENT
                                        if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                            ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                        else :
                                            ExistComment = ""
                                        ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in vlog")).font = Font(color='FF000000')
                                        #                                         
                                    ws.cell(row=rowNum,column=7,value="FAIL").font = Font(color='00FF0000')
                                    
                                    # COMMENT
                                    if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                        ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                    else :
                                        ExistComment = ""
                                    ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {vlogFile}\nERROR : Cell '{CellKey}' defined in vlog but not in lef")).font = Font(color='FF000000')                                            
                            #

    return wb,ws