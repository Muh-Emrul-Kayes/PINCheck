import os
import json
import sys
import re
from log import Logging
from utils import regexExtraction, write, jsonWrite,check_file, read,jsonRead, makeDirs,check_file
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment,PatternFill,Border,Side,Font 
from openpyxl.worksheet.table import Table, TableStyleInfo

# LIB Vs VLOG Compare
def compareLibVsVlog(wb,ws,libFileKeys, vlogFileKeys, libJson, vlogJson, IP_TYPE, ExcelKeys):
    """
    This function will generate the lib vs vlog comparison report as spreadsheet.

    Usage::
        >>> import libVsvlog
        >>> libVsvlog.compareLefVsVlog("WorkBook","WorkSheet","lib_file_list","vlog_file_list",
                                        "lib_json_data", "vlog_json_data", "IP_TYPE", "Excel_heading")

    :param: wb: workbook 
    :param: ws: worksheet
    :param: libFileKeys: lib file list 
    :param: vlogFileKeys: vlog file list
    :param: libJson: lib data as json dictionary format
    :param: vlogJson: vlog data as json dictionary format
    :param: IP_TYPE: ip type. Ex. memory, io, logic etc.
    :param: ExcelKeys: spreadsheet heading as list 
    :returns: wb: workbook
    :returns: ws: worksheet
    """  
    Logging.message("INFO", "GENERATING LIB Vs VERILOG COMPARISON REPORT")    
    for libFile in libFileKeys:
        for vlogFile in vlogFileKeys:
            libCellKeys = [*libJson[libFile].keys()]
            vlogCellKeys = [*vlogJson[vlogFile].keys()]
            MergeCellKeys = sorted(list(set(libCellKeys)|set(vlogCellKeys)))
            for CellKey in MergeCellKeys:
                if CellKey in libCellKeys and CellKey in vlogCellKeys:
                    libPortKeys = [*libJson[libFile][CellKey].keys()]
                    vlogPortKeys = [*vlogJson[vlogFile][CellKey].keys()]
                    MergePortKeys = sorted(list(set(libPortKeys)|set(vlogPortKeys)))
                    # IF PORT NOT FOUND IN LIB AND VLOG > ERROR COMMENT
                    if len(MergePortKeys) == 0:
                        #
                        FoundExistKey = False
                        for rowNum, rowWrite in enumerate(ws.iter_rows(values_only=True), start=1):
                            if rowWrite[0] == CellKey and rowWrite[1] == "NA":
                                # lib DIR
                                ws.cell(row=rowNum,column=4,value="NA").font = Font(color='000000FF')
                                # vlog DIR
                                ws.cell(row=rowNum,column=5,value="NA").font = Font(color='000000FF')
                                # ERROR COMMENT
                                if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                    ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                else :
                                    ExistComment = ""
                                Commentlib = f"ERROR : For Cell '{CellKey}' Pin not defined in lib"
                                commentvlog = f"ERROR : For Cell '{CellKey}' Pin not defined in vlog"
                                ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {libFile} Vs {vlogFile}\n{Commentlib}\n{commentvlog}")).font = Font(color='FF000000')
                                #                                                
                                # STATUS
                                ws.cell(row=rowNum,column=8,value="NA").font = Font(color='000000FF')
                                FoundExistKey = True
                                break                        
                        if not FoundExistKey:
                            FoundExistCell = False
                            for rowNum, rowWrite in enumerate(ws.iter_rows(values_only=True), start=1):
                                if rowWrite[0] == CellKey:
                                    ws.insert_rows(rowNum)
                                    ws.cell(row=rowNum,column=1,value=str(CellKey)).font = Font(color='FF000000')
                                    ws.cell(row=rowNum,column=2,value="NA").font = Font(color='000000FF')
                                    # lib DIR
                                    ws.cell(row=rowNum,column=4,value="NA").font = Font(color='000000FF')
                                    # vlog DIR
                                    ws.cell(row=rowNum,column=5,value="NA").font = Font(color='000000FF')
                                    # ERROR COMMENT
                                    if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                        ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                    else :
                                        ExistComment = ""
                                    Commentlib = f"ERROR : For Cell '{CellKey}' Pin not defined in lib"
                                    commentvlog = f"ERROR : For Cell '{CellKey}' Pin not defined in vlog"
                                    ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {libFile} Vs {vlogFile}\n{Commentlib}\n{commentvlog}")).font = Font(color='FF000000')
                                    #                                                 
                                    ws.cell(row=rowNum,column=8,value="NA").font = Font(color='000000FF')
                                    FoundExistCell = True
                                    break 
                            if not FoundExistCell:
                                rowNum = len(ws['A'])+1
                                ws.cell(row=rowNum,column=1,value=str(CellKey)).font = Font(color='FF000000')
                                ws.cell(row=rowNum,column=2,value="NA").font = Font(color='000000FF')
                                # lib DIR
                                ws.cell(row=rowNum,column=4,value="NA").font = Font(color='000000FF')
                                # vlog DIR
                                ws.cell(row=rowNum,column=5,value="NA").font = Font(color='000000FF')
                                # ERROR COMMENT
                                if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                    ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                else :
                                    ExistComment = ""
                                Commentlib = f"ERROR : For Cell '{CellKey}' Pin not defined in lib"
                                commentvlog = f"ERROR : For Cell '{CellKey}' Pin not defined in vlog"
                                ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {libFile} Vs {vlogFile}\n{Commentlib}\n{commentvlog}")).font = Font(color='FF000000')
                                #                                                 
                                ws.cell(row=rowNum,column=8,value="NA").font = Font(color='000000FF')
                        #
                    # Compare PIN between LIB and VLOG Cell
                    for PortKey in MergePortKeys:
                        if PortKey in libPortKeys and PortKey in vlogPortKeys:
                            libPortDir = libJson[libFile][CellKey][PortKey]["PortDirection"].upper()
                            vlogPortDir = vlogJson[vlogFile][CellKey][PortKey]["PortDirection"].upper()
                            if libPortDir == vlogPortDir and libPortDir.strip() and vlogPortDir.strip():
                                #
                                FoundExistKey = False
                                for rowNum, rowWrite in enumerate(ws.iter_rows(values_only=True), start=1):
                                    if rowWrite[0] == CellKey and rowWrite[1] == PortKey:
                                        # LIB DIR
                                        if (rowWrite[3] == None or rowWrite[3] == "NA") and rowWrite[3] != "ERROR":
                                            if libPortDir.strip():
                                                ws.cell(row=rowNum,column=4,value=str(libPortDir.upper())).font = Font(color='FF000000')
                                            else:
                                                ws.cell(row=rowNum,column=4,value="ERROR").font = Font(color='00FF0000')
                                                # ERROR COMMENT
                                                if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                    ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                                else :
                                                    ExistComment = ""
                                                ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {libFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in lib")).font = Font(color='FF000000')
                                                #
                                        else : 
                                            if rowWrite[3] == libPortDir.upper():
                                                pass
                                            else :
                                                ws.cell(row=rowNum,column=4,value="ERROR").font = Font(color='00FF0000')
                                                # ERROR COMMENT
                                                if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                    ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                                else :
                                                    ExistComment = ""
                                                if libPortDir.strip():
                                                    ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {libFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction mismatched within lib")).font = Font(color='FF000000')
                                                else:
                                                    ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {libFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in lib")).font = Font(color='FF000000')
                                                #
                                        # VLOG DIR
                                        if (rowWrite[4] == None or rowWrite[4] == "NA") and rowWrite[4] != "ERROR":
                                            if vlogPortDir.strip():
                                                ws.cell(row=rowNum,column=5,value=str(vlogPortDir.upper())).font = Font(color='FF000000')
                                            else:
                                                ws.cell(row=rowNum,column=5,value="ERROR").font = Font(color='00FF0000')
                                                # ERROR COMMENT
                                                if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                    ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                                else :
                                                    ExistComment = ""
                                                ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {libFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in vlog")).font = Font(color='FF000000')
                                                #                                                
                                        else : 
                                            if rowWrite[4] == vlogPortDir.upper():
                                                pass
                                            else :
                                                ws.cell(row=rowNum,column=5,value="ERROR").font = Font(color='00FF0000') 
                                                # ERROR COMMENT
                                                if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                    ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                                else :
                                                    ExistComment = ""
                                                if vlogPortDir.strip():
                                                    ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {libFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction mismatched within vlog")).font = Font(color='FF000000')
                                                else:
                                                    ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {libFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in vlog")).font = Font(color='FF000000')
                                                #                                                
                                        # STATUS
                                        if rowWrite[7] == None or rowWrite[7] == "NA":
                                            ws.cell(row=rowNum,column=8,value="PASS").font = Font(color='00008000')
                                        FoundExistKey = True
                                        break                        
                                if not FoundExistKey:
                                    FoundExistCell = False
                                    for rowNum, rowWrite in enumerate(ws.iter_rows(values_only=True), start=1):
                                        if rowWrite[0] == CellKey:
                                            ws.insert_rows(rowNum)
                                            ws.cell(row=rowNum,column=1,value=str(CellKey)).font = Font(color='FF000000')
                                            ws.cell(row=rowNum,column=2,value=str(PortKey)).font = Font(color='FF000000')
                                            if libPortDir.strip():
                                                ws.cell(row=rowNum,column=4,value=str(libPortDir.upper())).font = Font(color='FF000000')
                                            else:
                                                ws.cell(row=rowNum,column=4,value="ERROR").font = Font(color='00FF0000')
                                                # ERROR COMMENT
                                                if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                    ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                                else :
                                                    ExistComment = ""
                                                ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {libFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in lib")).font = Font(color='FF000000')
                                                #                                                 
                                            if vlogPortDir.strip():
                                                ws.cell(row=rowNum,column=5,value=str(vlogPortDir.upper())).font = Font(color='FF000000')
                                            else:
                                                ws.cell(row=rowNum,column=5,value="ERROR").font = Font(color='00FF0000')
                                                # ERROR COMMENT
                                                if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                    ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                                else :
                                                    ExistComment = ""
                                                ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {libFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in vlog")).font = Font(color='FF000000')
                                                #                                                 
                                            ws.cell(row=rowNum,column=8,value="PASS").font = Font(color='00008000')
                                            FoundExistCell = True
                                            break 
                                    if not FoundExistCell:
                                        rowNum = len(ws['A'])+1
                                        ws.cell(row=rowNum,column=1,value=str(CellKey)).font = Font(color='FF000000')
                                        ws.cell(row=rowNum,column=2,value=str(PortKey)).font = Font(color='FF000000')
                                        if libPortDir.strip():
                                            ws.cell(row=rowNum,column=4,value=str(libPortDir.upper())).font = Font(color='FF000000')
                                        else:
                                            ws.cell(row=rowNum,column=4,value="ERROR").font = Font(color='00FF0000')
                                            # ERROR COMMENT
                                            if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                            else :
                                                ExistComment = ""
                                            ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {libFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in lib")).font = Font(color='FF000000')
                                            #                                             
                                        if vlogPortDir.strip():
                                            ws.cell(row=rowNum,column=5,value=str(vlogPortDir.upper())).font = Font(color='FF000000')
                                        else:
                                            ws.cell(row=rowNum,column=5,value="ERROR").font = Font(color='00FF0000')
                                            # ERROR COMMENT
                                            if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                            else :
                                                ExistComment = ""
                                            ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {libFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in vlog")).font = Font(color='FF000000')
                                            #                                             
                                        ws.cell(row=rowNum,column=8,value="PASS").font = Font(color='00008000')
                                #
                                                        
                            else:
                                #
                                FoundExistKey = False
                                for rowNum, rowWrite in enumerate(ws.iter_rows(values_only=True), start=1):
                                    if rowWrite[0] == CellKey and rowWrite[1] == PortKey:
                                        # LIB DIR
                                        if (rowWrite[3] == None or rowWrite[3] == "NA") and rowWrite[3] != "ERROR":
                                            if libPortDir.strip():
                                                ws.cell(row=rowNum,column=4,value=str(libPortDir.upper())).font = Font(color='FF000000')
                                            else:
                                                ws.cell(row=rowNum,column=4,value="ERROR").font = Font(color='00FF0000')
                                                # ERROR COMMENT
                                                if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                    ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                                else :
                                                    ExistComment = ""
                                                ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {libFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in lib")).font = Font(color='FF000000')
                                                #                                                 
                                        else : 
                                            if rowWrite[3] == libPortDir.upper():
                                                pass
                                            else :
                                                ws.cell(row=rowNum,column=4,value="ERROR").font = Font(color='00FF0000')
                                                # ERROR COMMENT
                                                if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                    ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                                else :
                                                    ExistComment = ""
                                                if libPortDir.strip():
                                                    ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {libFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction mismatched within lib")).font = Font(color='FF000000')
                                                else:
                                                    ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {libFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in lib")).font = Font(color='FF000000')
                                                #                                                
                                        # VLOG DIR
                                        if (rowWrite[4] == None or rowWrite[4] == "NA") and rowWrite[4] != "ERROR":
                                            if vlogPortDir.strip():
                                                ws.cell(row=rowNum,column=5,value=str(vlogPortDir.upper())).font = Font(color='FF000000')
                                            else:
                                                ws.cell(row=rowNum,column=5,value="ERROR").font = Font(color='00FF0000')
                                                # ERROR COMMENT
                                                if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                    ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                                else :
                                                    ExistComment = ""
                                                ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {libFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in vlog")).font = Font(color='FF000000')
                                                #                                                 
                                        else : 
                                            if rowWrite[4] == vlogPortDir.upper():
                                                pass
                                            else :
                                                ws.cell(row=rowNum,column=5,value="ERROR").font = Font(color='00FF0000')
                                                # ERROR COMMENT
                                                if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                    ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                                else :
                                                    ExistComment = ""
                                                if vlogPortDir.strip():
                                                    ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {libFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction mismatched within vlog")).font = Font(color='FF000000')
                                                else:
                                                    ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {libFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in vlog")).font = Font(color='FF000000')
                                                #                                                 
                                        # STATUS
                                        if rowWrite[7] == None or rowWrite[7] == "NA" or rowWrite[7] == "PASS":
                                            ws.cell(row=rowNum,column=8,value="ERROR").font = Font(color='00FF0000')
                                        # COMMENT
                                        if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                            ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                        else :
                                            ExistComment = ""
                                        ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {libFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction mismatched between lib and vlog")).font = Font(color='FF000000')                                                                                        
                                        FoundExistKey = True
                                        break                        
                                if not FoundExistKey:
                                    FoundExistCell = False
                                    for rowNum, rowWrite in enumerate(ws.iter_rows(values_only=True), start=1):
                                        if rowWrite[0] == CellKey:
                                            ws.insert_rows(rowNum)
                                            ws.cell(row=rowNum,column=1,value=str(CellKey)).font = Font(color='FF000000')
                                            ws.cell(row=rowNum,column=2,value=str(PortKey)).font = Font(color='FF000000')
                                            if libPortDir.strip():
                                                ws.cell(row=rowNum,column=4,value=str(libPortDir.upper())).font = Font(color='FF000000')
                                            else:
                                                ws.cell(row=rowNum,column=4,value="ERROR").font = Font(color='00FF0000')
                                                # ERROR COMMENT
                                                if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                    ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                                else :
                                                    ExistComment = ""
                                                ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {libFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in lib")).font = Font(color='FF000000')
                                                #                                                 
                                            if vlogPortDir.strip():
                                                ws.cell(row=rowNum,column=5,value=str(vlogPortDir.upper())).font = Font(color='FF000000')
                                            else:
                                                ws.cell(row=rowNum,column=5,value="ERROR").font = Font(color='00FF0000')
                                                # ERROR COMMENT
                                                if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                    ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                                else :
                                                    ExistComment = ""
                                                ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {libFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in vlog")).font = Font(color='FF000000')
                                                #                                                 
                                            ws.cell(row=rowNum,column=8,value="ERROR").font = Font(color='00FF0000')
                                            # COMMENT
                                            if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                            else :
                                                ExistComment = ""
                                            ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {libFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction mismatched between lib and vlog")).font = Font(color='FF000000')                                            
                                            FoundExistCell = True
                                            break 
                                    if not FoundExistCell:
                                        rowNum = len(ws['A'])+1
                                        ws.cell(row=rowNum,column=1,value=str(CellKey)).font = Font(color='FF000000')
                                        ws.cell(row=rowNum,column=2,value=str(PortKey)).font = Font(color='FF000000')
                                        if libPortDir.strip():
                                            ws.cell(row=rowNum,column=4,value=str(libPortDir.upper())).font = Font(color='FF000000')
                                        else:
                                            ws.cell(row=rowNum,column=4,value="ERROR").font = Font(color='00FF0000')
                                            # ERROR COMMENT
                                            if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                            else :
                                                ExistComment = ""
                                            ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {libFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in lib")).font = Font(color='FF000000')
                                            #                                             
                                        if vlogPortDir.strip():
                                            ws.cell(row=rowNum,column=5,value=str(vlogPortDir.upper())).font = Font(color='FF000000')
                                        else:
                                            ws.cell(row=rowNum,column=5,value="ERROR").font = Font(color='00FF0000')
                                            # ERROR COMMENT
                                            if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                            else :
                                                ExistComment = ""
                                            ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {libFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in vlog")).font = Font(color='FF000000')
                                            #                                             
                                        ws.cell(row=rowNum,column=8,value="ERROR").font = Font(color='00FF0000')
                                        # COMMENT
                                        if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                            ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                        else :
                                            ExistComment = ""
                                        ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {libFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction mismatched between lib and vlog")).font = Font(color='FF000000')                                         
                                #                                
                        # FOR PORT NOT FOUND IN CELL BETWEEN LIB AND VLOG
                        else:
                            if PortKey in libPortKeys:
                                libPortDir = libJson[libFile][CellKey][PortKey]["PortDirection"].upper()  
                                #
                                FoundExistKey = False
                                for rowNum, rowWrite in enumerate(ws.iter_rows(values_only=True), start=1):
                                    if rowWrite[0] == CellKey and rowWrite[1] == PortKey:
                                        # LIB DIR
                                        if (rowWrite[3] == None or rowWrite[3] == "NA") and rowWrite[3] != "ERROR":
                                            if libPortDir.strip():
                                                ws.cell(row=rowNum,column=4,value=str(libPortDir.upper())).font = Font(color='FF000000')
                                            else:
                                                ws.cell(row=rowNum,column=4,value="ERROR").font = Font(color='00FF0000')
                                                # ERROR COMMENT
                                                if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                    ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                                else :
                                                    ExistComment = ""
                                                ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {libFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in lib")).font = Font(color='FF000000')
                                                #                                                 
                                        else : 
                                            if rowWrite[3] == libPortDir.upper():
                                                pass
                                            else :
                                                ws.cell(row=rowNum,column=4,value="ERROR").font = Font(color='00FF0000')
                                                # ERROR COMMENT
                                                if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                    ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                                else :
                                                    ExistComment = ""
                                                if libPortDir.strip():
                                                    ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {libFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction mismatched within lib")).font = Font(color='FF000000')
                                                else:
                                                    ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {libFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in lib")).font = Font(color='FF000000')
                                                #                                                
                                        # VLOG DIR
                                        if rowWrite[4] == None:
                                            ws.cell(row=rowNum,column=5,value="NA").font = Font(color='000000FF') 
                                        # STATUS
                                        # if rowWrite[5] != "PASS":
                                        #     ws.cell(row=rowNum,column=6,value="FAIL").font = Font(color='00FF0000')
                                        
                                        ws.cell(row=rowNum,column=8,value="FAIL").font = Font(color='00FF0000')
                                                                                   
                                        # COMMENT
                                        if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                            ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                        else :
                                            ExistComment = ""
                                        ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {libFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' defined in lib but not in vlog")).font = Font(color='FF000000')
                                        FoundExistKey = True
                                        break                        
                                if not FoundExistKey:
                                    FoundExistCell = False
                                    for rowNum, rowWrite in enumerate(ws.iter_rows(values_only=True), start=1):
                                        if rowWrite[0] == CellKey:
                                            ws.insert_rows(rowNum)
                                            ws.cell(row=rowNum,column=1,value=str(CellKey)).font = Font(color='FF000000')
                                            ws.cell(row=rowNum,column=2,value=str(PortKey)).font = Font(color='FF000000')
                                            if libPortDir.strip():
                                                ws.cell(row=rowNum,column=4,value=str(libPortDir.upper())).font = Font(color='FF000000')
                                            else:
                                                ws.cell(row=rowNum,column=4,value="ERROR").font = Font(color='00FF0000')
                                                # ERROR COMMENT
                                                if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                    ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                                else :
                                                    ExistComment = ""
                                                ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {libFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in lib")).font = Font(color='FF000000')
                                                #                                                 
                                            ws.cell(row=rowNum,column=5,value="NA").font = Font(color='000000FF')
                                            ws.cell(row=rowNum,column=8,value="FAIL").font = Font(color='00FF0000')
                                            
                                            # COMMENT
                                            if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                            else :
                                                ExistComment = ""
                                            ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {libFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' defined in lib but not in vlog")).font = Font(color='FF000000')                                            
                                            FoundExistCell = True
                                            break 
                                    if not FoundExistCell:
                                        rowNum = len(ws['A'])+1
                                        ws.cell(row=rowNum,column=1,value=str(CellKey)).font = Font(color='FF000000')
                                        ws.cell(row=rowNum,column=2,value=str(PortKey)).font = Font(color='FF000000')
                                        if libPortDir.strip():
                                            ws.cell(row=rowNum,column=4,value=str(libPortDir.upper())).font = Font(color='FF000000')
                                        else:
                                            ws.cell(row=rowNum,column=4,value="ERROR").font = Font(color='00FF0000')
                                            # ERROR COMMENT
                                            if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                            else :
                                                ExistComment = ""
                                            ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {libFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in lib")).font = Font(color='FF000000')
                                            #                                             
                                        ws.cell(row=rowNum,column=5,value="NA").font = Font(color='000000FF')
                                        ws.cell(row=rowNum,column=8,value="FAIL").font = Font(color='00FF0000')
                                        
                                        # COMMENT
                                        if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                            ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                        else :
                                            ExistComment = ""
                                        ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {libFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' defined in lib but not in vlog")).font = Font(color='FF000000')                                            
                                #                                   
                            else:   
                                vlogPortDir = vlogJson[vlogFile][CellKey][PortKey]["PortDirection"].upper()   
                                #
                                FoundExistKey = False
                                for rowNum, rowWrite in enumerate(ws.iter_rows(values_only=True), start=1):
                                    if rowWrite[0] == CellKey and rowWrite[1] == PortKey:
                                        # lib DIR
                                        if rowWrite[3] == None:
                                            ws.cell(row=rowNum,column=4,value="NA").font = Font(color='000000FF')
                                        # VLOG DIR 
                                        if (rowWrite[4] == None or rowWrite[4] == "NA") and rowWrite[4] != "ERROR":
                                            if vlogPortDir.strip():
                                                ws.cell(row=rowNum,column=5,value=str(vlogPortDir.upper())).font = Font(color='FF000000')
                                            else:
                                                ws.cell(row=rowNum,column=5,value="ERROR").font = Font(color='00FF0000')
                                                # ERROR COMMENT
                                                if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                    ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                                else :
                                                    ExistComment = ""
                                                ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {libFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in vlog")).font = Font(color='FF000000')
                                                #                                                 
                                        else : 
                                            if rowWrite[4] == vlogPortDir.upper():
                                                pass
                                            else :
                                                ws.cell(row=rowNum,column=5,value="ERROR").font = Font(color='00FF0000')         
                                                # ERROR COMMENT
                                                if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                    ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                                else :
                                                    ExistComment = ""
                                                if vlogPortDir.strip():
                                                    ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {libFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction mismatched within vlog")).font = Font(color='FF000000')
                                                else:
                                                    ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {libFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in vlog")).font = Font(color='FF000000')
                                                #                                                                                   
                                        # STATUS
                                        # if rowWrite[7] != "PASS":
                                        #     ws.cell(row=rowNum,column=8,value="FAIL").font = Font(color='00FF0000')
                                        
                                        ws.cell(row=rowNum,column=8,value="FAIL").font = Font(color='00FF0000')
                                        
                                        # COMMENT
                                        if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                            ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                        else :
                                            ExistComment = ""
                                        ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {libFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' defined in vlog but not in lib")).font = Font(color='FF000000')
                                        FoundExistKey = True
                                        break                        
                                if not FoundExistKey:
                                    FoundExistCell = False
                                    for rowNum, rowWrite in enumerate(ws.iter_rows(values_only=True), start=1):
                                        if rowWrite[0] == CellKey:
                                            ws.insert_rows(rowNum)
                                            ws.cell(row=rowNum,column=1,value=str(CellKey)).font = Font(color='FF000000')
                                            ws.cell(row=rowNum,column=2,value=str(PortKey)).font = Font(color='FF000000')
                                            ws.cell(row=rowNum,column=4,value="NA").font = Font(color='000000FF')
                                            if vlogPortDir.strip():
                                                ws.cell(row=rowNum,column=5,value=str(vlogPortDir.upper())).font = Font(color='FF000000')
                                            else:
                                                ws.cell(row=rowNum,column=5,value="ERROR").font = Font(color='00FF0000')
                                                # ERROR COMMENT
                                                if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                    ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                                else :
                                                    ExistComment = ""
                                                ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {libFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in vlog")).font = Font(color='FF000000')
                                                #                                                 
                                            ws.cell(row=rowNum,column=8,value="FAIL").font = Font(color='00FF0000')
                                            
                                            # COMMENT
                                            if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                            else :
                                                ExistComment = ""
                                            ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {libFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' defined in vlog but not in lib")).font = Font(color='FF000000')                                            
                                            FoundExistCell = True
                                            break 
                                    if not FoundExistCell:
                                        rowNum = len(ws['A'])+1
                                        ws.cell(row=rowNum,column=1,value=str(CellKey)).font = Font(color='FF000000')
                                        ws.cell(row=rowNum,column=2,value=str(PortKey)).font = Font(color='FF000000')
                                        ws.cell(row=rowNum,column=4,value="NA").font = Font(color='000000FF')
                                        if vlogPortDir.strip():
                                            ws.cell(row=rowNum,column=5,value=str(vlogPortDir.upper())).font = Font(color='FF000000')
                                        else:
                                            ws.cell(row=rowNum,column=5,value="ERROR").font = Font(color='00FF0000')
                                            # ERROR COMMENT
                                            if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                            else :
                                                ExistComment = ""
                                            ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {libFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in vlog")).font = Font(color='FF000000')
                                            #                                             
                                        ws.cell(row=rowNum,column=8,value="FAIL").font = Font(color='00FF0000')
                                        
                                        # COMMENT
                                        if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                            ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                        else :
                                            ExistComment = ""
                                        ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {libFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' defined in vlog but not in lib")).font = Font(color='FF000000')                                            
                                #                                       
                # FOR CELL NOT FOUND IN ONE FILE LIB OR VLOG        
                else:
                    if CellKey in libCellKeys:
                        libPortKeys = [*libJson[libFile][CellKey].keys()]
                        # IF PORT NOT FOUND IN LIB > ERROR COMMENT
                        if len(libPortKeys) == 0:
                            #
                            FoundExistKey = False
                            for rowNum, rowWrite in enumerate(ws.iter_rows(values_only=True), start=1):
                                if rowWrite[0] == CellKey and rowWrite[1] == "NA":
                                    # lib DIR
                                    ws.cell(row=rowNum,column=4,value="NA").font = Font(color='000000FF')
                                    # vlog DIR
                                    ws.cell(row=rowNum,column=5,value="NA").font = Font(color='000000FF')
                                    # ERROR COMMENT
                                    if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                        ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                    else :
                                        ExistComment = ""
                                    Commentlib = f"ERROR : For Cell '{CellKey}' Pin not defined in lib"
                                    commentvlog = f"ERROR : Cell '{CellKey}' not defined in vlog"
                                    ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {libFile} Vs {vlogFile}\n{Commentlib}\n{commentvlog}")).font = Font(color='FF000000')
                                    #                                                
                                    # STATUS
                                    ws.cell(row=rowNum,column=8,value="FAIL").font = Font(color='00FF0000')
                                    FoundExistKey = True
                                    break                        
                            if not FoundExistKey:
                                FoundExistCell = False
                                for rowNum, rowWrite in enumerate(ws.iter_rows(values_only=True), start=1):
                                    if rowWrite[0] == CellKey:
                                        ws.insert_rows(rowNum)
                                        ws.cell(row=rowNum,column=1,value=str(CellKey)).font = Font(color='FF000000')
                                        ws.cell(row=rowNum,column=2,value="NA").font = Font(color='000000FF')
                                        # lib DIR
                                        ws.cell(row=rowNum,column=4,value="NA").font = Font(color='000000FF')
                                        # vlog DIR
                                        ws.cell(row=rowNum,column=5,value="NA").font = Font(color='000000FF')
                                        # ERROR COMMENT
                                        if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                            ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                        else :
                                            ExistComment = ""
                                        Commentlib = f"ERROR : For Cell '{CellKey}' Pin not defined in lib"
                                        commentvlog = f"ERROR : Cell '{CellKey}' not defined in vlog"
                                        ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {libFile} Vs {vlogFile}\n{Commentlib}\n{commentvlog}")).font = Font(color='FF000000')
                                        #                                                
                                        # STATUS
                                        ws.cell(row=rowNum,column=8,value="FAIL").font = Font(color='00FF0000')
                                        FoundExistCell = True
                                        break 
                                if not FoundExistCell:
                                    rowNum = len(ws['A'])+1
                                    ws.cell(row=rowNum,column=1,value=str(CellKey)).font = Font(color='FF000000')
                                    ws.cell(row=rowNum,column=2,value="NA").font = Font(color='000000FF')
                                    # lib DIR
                                    ws.cell(row=rowNum,column=4,value="NA").font = Font(color='000000FF')
                                    # vlog DIR
                                    ws.cell(row=rowNum,column=5,value="NA").font = Font(color='000000FF')
                                    # ERROR COMMENT
                                    if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                        ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                    else :
                                        ExistComment = ""
                                    Commentlib = f"ERROR : For Cell '{CellKey}' Pin not defined in lib"
                                    commentvlog = f"ERROR : Cell '{CellKey}' not defined in vlog"
                                    ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {libFile} Vs {vlogFile}\n{Commentlib}\n{commentvlog}")).font = Font(color='FF000000')
                                    #                                                
                                    # STATUS
                                    ws.cell(row=rowNum,column=8,value="FAIL").font = Font(color='00FF0000')
                            #                        
                        for PortKey in libPortKeys:
                            libPortDir = libJson[libFile][CellKey][PortKey]["PortDirection"].upper()   
                            #
                            FoundExistKey = False
                            for rowNum, rowWrite in enumerate(ws.iter_rows(values_only=True), start=1):
                                if rowWrite[0] == CellKey and rowWrite[1] == PortKey:
                                    # lib DIR
                                    if (rowWrite[3] == None or rowWrite[3] == "NA") and rowWrite[3] != "ERROR":
                                        if libPortDir.strip():
                                            ws.cell(row=rowNum,column=4,value=str(libPortDir.upper())).font = Font(color='FF000000')
                                        else:
                                            ws.cell(row=rowNum,column=4,value="ERROR").font = Font(color='00FF0000')
                                            # ERROR COMMENT
                                            if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                            else :
                                                ExistComment = ""
                                            ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {libFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in lib")).font = Font(color='FF000000')
                                            #                                             
                                    else : 
                                        if rowWrite[3] == libPortDir.upper():
                                            pass
                                        else :
                                            ws.cell(row=rowNum,column=4,value="ERROR").font = Font(color='00FF0000')  
                                            # ERROR COMMENT
                                            if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                            else :
                                                ExistComment = ""
                                            if libPortDir.strip():
                                                ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {libFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction mismatched within lib")).font = Font(color='FF000000')
                                            else:
                                                ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {libFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in lib")).font = Font(color='FF000000')
                                            #                                                                               
                                    # VLOG DIR 
                                    if rowWrite[4] == None:
                                        ws.cell(row=rowNum,column=5,value="NA").font = Font(color='000000FF')                                             
                                    # STATUS
                                    # if rowWrite[7] != "PASS":
                                    #     ws.cell(row=rowNum,column=8,value="FAIL").font = Font(color='00FF0000')
                                    
                                    ws.cell(row=rowNum,column=8,value="FAIL").font = Font(color='00FF0000')  
                                    
                                    # COMMENT
                                    if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                        ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                    else :
                                        ExistComment = ""
                                    ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {libFile} Vs {vlogFile}\nERROR : Cell '{CellKey}' defined in lib but not in vlog")).font = Font(color='FF000000')
                                    FoundExistKey = True
                                    break                        
                            if not FoundExistKey:
                                FoundExistCell = False
                                for rowNum, rowWrite in enumerate(ws.iter_rows(values_only=True), start=1):
                                    if rowWrite[0] == CellKey:
                                        ws.insert_rows(rowNum)
                                        ws.cell(row=rowNum,column=1,value=str(CellKey)).font = Font(color='FF000000')
                                        ws.cell(row=rowNum,column=2,value=str(PortKey)).font = Font(color='FF000000')
                                        if libPortDir.strip():
                                            ws.cell(row=rowNum,column=4,value=str(libPortDir.upper())).font = Font(color='FF000000')
                                        else:
                                            ws.cell(row=rowNum,column=4,value="ERROR").font = Font(color='00FF0000')
                                            # ERROR COMMENT
                                            if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                            else :
                                                ExistComment = ""
                                            ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {libFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in lib")).font = Font(color='FF000000')
                                            #                                            
                                        ws.cell(row=rowNum,column=5,value="NA").font = Font(color='000000FF')
                                        ws.cell(row=rowNum,column=8,value="FAIL").font = Font(color='00FF0000')
                                        
                                        # COMMENT
                                        if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                            ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                        else :
                                            ExistComment = ""
                                        ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {libFile} Vs {vlogFile}\nERROR : Cell '{CellKey}' defined in lib but not in vlog")).font = Font(color='FF000000')                                            
                                        FoundExistCell = True
                                        break 
                                if not FoundExistCell:
                                    rowNum = len(ws['A'])+1
                                    ws.cell(row=rowNum,column=1,value=str(CellKey)).font = Font(color='FF000000')
                                    ws.cell(row=rowNum,column=2,value=str(PortKey)).font = Font(color='FF000000')
                                    if libPortDir.strip():
                                        ws.cell(row=rowNum,column=4,value=str(libPortDir.upper())).font = Font(color='FF000000')
                                    else:
                                        ws.cell(row=rowNum,column=4,value="ERROR").font = Font(color='00FF0000')
                                        # ERROR COMMENT
                                        if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                            ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                        else :
                                            ExistComment = ""
                                        ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {libFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in lib")).font = Font(color='FF000000')
                                        #                                        
                                    ws.cell(row=rowNum,column=5,value="NA").font = Font(color='000000FF')
                                    ws.cell(row=rowNum,column=8,value="FAIL").font = Font(color='00FF0000')
                                    
                                    # COMMENT
                                    if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                        ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                    else :
                                        ExistComment = ""
                                    ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {libFile} Vs {vlogFile}\nERROR : Cell '{CellKey}' defined in lib but not in vlog")).font = Font(color='FF000000')                                            
                            #                        
                    else: 
                        vlogPortKeys = [*vlogJson[vlogFile][CellKey].keys()]
                        # IF PORT NOT FOUND IN VLOG > ERROR COMMENT
                        if len(vlogPortKeys) == 0:
                            #
                            FoundExistKey = False
                            for rowNum, rowWrite in enumerate(ws.iter_rows(values_only=True), start=1):
                                if rowWrite[0] == CellKey and rowWrite[1] == "NA":
                                    # lib DIR
                                    ws.cell(row=rowNum,column=4,value="NA").font = Font(color='000000FF')
                                    # vlog DIR
                                    ws.cell(row=rowNum,column=5,value="NA").font = Font(color='000000FF')
                                    # ERROR COMMENT
                                    if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                        ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                    else :
                                        ExistComment = ""
                                    Commentlib = f"ERROR : Cell '{CellKey}' not defined in lib"
                                    commentvlog = f"ERROR : For Cell '{CellKey}' Pin not defined in vlog"
                                    ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {libFile} Vs {vlogFile}\n{Commentlib}\n{commentvlog}")).font = Font(color='FF000000')
                                    #                                                
                                    # STATUS
                                    ws.cell(row=rowNum,column=8,value="FAIL").font = Font(color='00FF0000')
                                    FoundExistKey = True
                                    break                        
                            if not FoundExistKey:
                                FoundExistCell = False
                                for rowNum, rowWrite in enumerate(ws.iter_rows(values_only=True), start=1):
                                    if rowWrite[0] == CellKey:
                                        ws.insert_rows(rowNum)
                                        ws.cell(row=rowNum,column=1,value=str(CellKey)).font = Font(color='FF000000')
                                        ws.cell(row=rowNum,column=2,value="NA").font = Font(color='000000FF')
                                        # lib DIR
                                        ws.cell(row=rowNum,column=4,value="NA").font = Font(color='000000FF')
                                        # vlog DIR
                                        ws.cell(row=rowNum,column=5,value="NA").font = Font(color='000000FF')
                                        # ERROR COMMENT
                                        if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                            ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                        else :
                                            ExistComment = ""
                                        Commentlib = f"ERROR : Cell '{CellKey}' not defined in lib"
                                        commentvlog = f"ERROR : For Cell '{CellKey}' Pin not defined in vlog"
                                        ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {libFile} Vs {vlogFile}\n{Commentlib}\n{commentvlog}")).font = Font(color='FF000000')
                                        #                                                
                                        # STATUS
                                        ws.cell(row=rowNum,column=8,value="FAIL").font = Font(color='00FF0000')
                                        FoundExistCell = True
                                        break 
                                if not FoundExistCell:
                                    rowNum = len(ws['A'])+1
                                    ws.cell(row=rowNum,column=1,value=str(CellKey)).font = Font(color='FF000000')
                                    ws.cell(row=rowNum,column=2,value="NA").font = Font(color='000000FF')
                                    # lib DIR
                                    ws.cell(row=rowNum,column=4,value="NA").font = Font(color='000000FF')
                                    # vlog DIR
                                    ws.cell(row=rowNum,column=5,value="NA").font = Font(color='000000FF')
                                    # ERROR COMMENT
                                    if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                        ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                    else :
                                        ExistComment = ""
                                    Commentlib = f"ERROR : Cell '{CellKey}' not defined in lib"
                                    commentvlog = f"ERROR : For Cell '{CellKey}' Pin not defined in vlog"
                                    ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {libFile} Vs {vlogFile}\n{Commentlib}\n{commentvlog}")).font = Font(color='FF000000')
                                    #                                                
                                    # STATUS
                                    ws.cell(row=rowNum,column=8,value="FAIL").font = Font(color='00FF0000')
                            #                       
                        for PortKey in vlogPortKeys:
                            vlogPortDir = vlogJson[vlogFile][CellKey][PortKey]["PortDirection"].upper()   
                            #
                            FoundExistKey = False
                            for rowNum, rowWrite in enumerate(ws.iter_rows(values_only=True), start=1):
                                if rowWrite[0] == CellKey and rowWrite[1] == PortKey:
                                    # lib DIR
                                    if rowWrite[3] == None:
                                        ws.cell(row=rowNum,column=4,value="NA").font = Font(color='000000FF')                                   
                                    # VLOG DIR 
                                    if (rowWrite[4] == None or rowWrite[4] == "NA") and rowWrite[4] != "ERROR":
                                        if vlogPortDir.strip():
                                            ws.cell(row=rowNum,column=5,value=str(vlogPortDir.upper())).font = Font(color='FF000000')
                                        else:
                                            ws.cell(row=rowNum,column=5,value="ERROR").font = Font(color='00FF0000')
                                            # ERROR COMMENT
                                            if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                            else :
                                                ExistComment = ""
                                            ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {libFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in vlog")).font = Font(color='FF000000')
                                            #                                            
                                    else : 
                                        if rowWrite[4] == vlogPortDir.upper():
                                            pass
                                        else :
                                            ws.cell(row=rowNum,column=5,value="ERROR").font = Font(color='00FF0000')
                                            # ERROR COMMENT
                                            if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                            else :
                                                ExistComment = ""
                                            if vlogPortDir.strip():
                                                ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {libFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction mismatched within vlog")).font = Font(color='FF000000')
                                            else:
                                                ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {libFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in vlog")).font = Font(color='FF000000')
                                            #
                                    # STATUS
                                    # if rowWrite[7] != "PASS":
                                    #     ws.cell(row=rowNum,column=8,value="FAIL").font = Font(color='00FF0000')
                                    
                                    ws.cell(row=rowNum,column=8,value="FAIL").font = Font(color='00FF0000')  
                                    
                                    # COMMENT
                                    if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                        ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                    else :
                                        ExistComment = ""
                                    ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {libFile} Vs {vlogFile}\nERROR : Cell '{CellKey}' defined in vlog but not in lib")).font = Font(color='FF000000')
                                    FoundExistKey = True
                                    break                        
                            if not FoundExistKey:
                                FoundExistCell = False
                                for rowNum, rowWrite in enumerate(ws.iter_rows(values_only=True), start=1):
                                    if rowWrite[0] == CellKey:
                                        ws.insert_rows(rowNum)
                                        ws.cell(row=rowNum,column=1,value=str(CellKey)).font = Font(color='FF000000')
                                        ws.cell(row=rowNum,column=2,value=str(PortKey)).font = Font(color='FF000000')
                                        ws.cell(row=rowNum,column=4,value="NA").font = Font(color='000000FF')
                                        if vlogPortDir.strip():
                                            ws.cell(row=rowNum,column=5,value=str(vlogPortDir.upper())).font = Font(color='FF000000')
                                        else:
                                            ws.cell(row=rowNum,column=5,value="ERROR").font = Font(color='00FF0000')
                                            # ERROR COMMENT
                                            if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                            else :
                                                ExistComment = ""
                                            ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {libFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in vlog")).font = Font(color='FF000000')
                                            #                                             
                                        ws.cell(row=rowNum,column=8,value="FAIL").font = Font(color='00FF0000')
                                        
                                        # COMMENT
                                        if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                            ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                        else :
                                            ExistComment = ""
                                        ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {libFile} Vs {vlogFile}\nERROR : Cell '{CellKey}' defined in vlog but not in lib")).font = Font(color='FF000000')                                            
                                        FoundExistCell = True
                                        break 
                                if not FoundExistCell:
                                    rowNum = len(ws['A'])+1
                                    ws.cell(row=rowNum,column=1,value=str(CellKey)).font = Font(color='FF000000')
                                    ws.cell(row=rowNum,column=2,value=str(PortKey)).font = Font(color='FF000000')
                                    ws.cell(row=rowNum,column=4,value="NA").font = Font(color='000000FF')
                                    if vlogPortDir.strip():
                                        ws.cell(row=rowNum,column=5,value=str(vlogPortDir.upper())).font = Font(color='FF000000')
                                    else:
                                        ws.cell(row=rowNum,column=5,value="ERROR").font = Font(color='00FF0000')
                                        # ERROR COMMENT
                                        if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                            ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                        else :
                                            ExistComment = ""
                                        ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {libFile} Vs {vlogFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in vlog")).font = Font(color='FF000000')
                                        #                                         
                                    ws.cell(row=rowNum,column=8,value="FAIL").font = Font(color='00FF0000')
                                    
                                    # COMMENT
                                    if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                        ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                    else :
                                        ExistComment = ""
                                    ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {libFile} Vs {vlogFile}\nERROR : Cell '{CellKey}' defined in vlog but not in lib")).font = Font(color='FF000000')                                            
                            #

    return wb,ws