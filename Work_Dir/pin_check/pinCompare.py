import os
import json
import sys
import re
from log import Logging
from utils import regexExtraction, write, jsonWrite,check_file, read,jsonRead, makeDirs,check_file
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment,PatternFill,Border,Side,Font 
from openpyxl.worksheet.table import Table, TableStyleInfo
from lefVslib import compareLefVsLib
from lefVsvlog import compareLefVsVlog
from libVsvlog import compareLibVsVlog

# Open Excel Workbook
def Open_Workbook(SheetName):
    """
    This function will Open a spreadsheet file for data extraction.

    Usage::
        >>> import pinCompare
        >>> pinCompare.Open_Workbook("WORK_SHEET_NAME")

    :param: SheetName: work sheet name
    :returns: wb: workbook
    :returns: ws: worksheet
    """
    wb = Workbook()
    ws = wb.active
    ws.title = SheetName
    return wb, ws

# Insert Headings 
def insertExcelKeys(ExcelKeys,ws):
    """
    This function will insert the column heading into spreadsheet file.

    Usage::
        >>> import pinCompare
        >>> pinCompare.insertExcelKeys("HEADING_LIST","WORK_SHEET")

    :param: ExcelKeys: heading as list
    :param: ws: worksheet 
    :returns: ws: worksheet
    """   
    bold_regular_font = Font(name='Calibri', size=11, bold=True, italic=False,
                                vertAlign=None, underline='none', strike=False, color='FF000000')
    alignment = Alignment(horizontal='center', vertical='center',
                            text_rotation=0, wrap_text=False, shrink_to_fit=False, indent=0)
    pattern_fill = PatternFill(start_color='00FFFF00',
                                end_color='00FFFF00', fill_type='solid')
    thin_border = Border(left=Side(border_style='thin', color='FF000000'),
                            right=Side(border_style='thin', color='FF000000'),
                            top=Side(border_style='thin', color='FF000000'),
                            bottom=Side(border_style='thin',
                                        color='FF000000'),
                            )
    for Num,Name in enumerate(ExcelKeys, start=1):
        cell_num = ws.cell(1,Num)
        cell_num.value = ExcelKeys[Num-1]
        if ExcelKeys[Num-1] == "Cell Name" :
            ws.column_dimensions[cell_num.column_letter].width = 35
        elif ExcelKeys[Num-1] == "Comments":
            ws.column_dimensions[cell_num.column_letter].width = 80
        else:
            ws.column_dimensions[cell_num.column_letter].width = 18

        ws.cell(1,Num).alignment = alignment
        ws.cell(1,Num).font = bold_regular_font
        ws.cell(1,Num).border = thin_border
    return ws

# PIN COMPARE BETWEEN VIEWS
def compare(lefJson, libJson, vlogJson, IP_TYPE):
    """
    This function will do comparison between lef,lib and verilog files.
    """
    wb, ws = Open_Workbook("FECheckInfo")
    ExcelKeys = ["Cell Name", "Pin Name", "Pin Direction LEF", "Pin Direction LIB", "Pin Direction VLOG", 
                    "LEF Vs LIB Status", "LEF Vs VLOG Status", "LIB Vs VLOG Status", "Comments"]
    ws = insertExcelKeys(ExcelKeys,ws)    

    if lefJson and libJson:
        lefFileKeys = [*lefJson.keys()]
        libFileKeys = [*libJson.keys()]
        wb,ws = compareLefVsLib(wb,ws,lefFileKeys, libFileKeys, lefJson, libJson, IP_TYPE, ExcelKeys)
    else:
        Logging.message("WARNING", f"LEF VS LIB COMPARISON IGNORED")

    if lefJson and vlogJson:
        lefFileKeys = [*lefJson.keys()]
        vlogFileKeys = [*vlogJson.keys()]
        wb,ws = compareLefVsVlog(wb,ws,lefFileKeys, vlogFileKeys, lefJson, vlogJson, IP_TYPE, ExcelKeys)
    else:
        Logging.message("WARNING", f"LEF VS VLOG COMPARISON IGNORED")

    if libJson and vlogJson:
        libFileKeys = [*libJson.keys()]
        vlogFileKeys = [*vlogJson.keys()]        
        wb,ws = compareLibVsVlog(wb,ws,libFileKeys, vlogFileKeys, libJson, vlogJson, IP_TYPE, ExcelKeys)
    else:
        Logging.message("WARNING", f"LIB VS VLOG COMPARISON IGNORED")
            
    ws = fllRows(ws,ExcelKeys)
    return wb,ws
    
# Style Rows
def fllRows(ws, ExcelKeys):
    """
    This function will insert the column colour and table format into spreadsheet file.

    Usage::
        >>> import pinCompare
        >>> pinCompare.fllRows("WORK_SHEET","HEADING_LIST")

    :param: ExcelKeys: heading as list
    :param: ws: worksheet 
    :returns: ws: worksheet
    """ 
    thin_border = Border(left=Side(border_style='thin', color='FF000000'),
                            right=Side(border_style='thin', color='FF000000'),
                            top=Side(border_style='thin', color='FF000000'),
                            bottom=Side(border_style='thin',
                                        color='FF000000'),
                            )
    for rowNum, rowWrite in enumerate(ws.iter_rows(values_only=True), start=1):
            if rowWrite[2] == None:
                ws.cell(row=rowNum,column=3,value="NA").font = Font(color='000000FF')
            if rowWrite[3] == None:
                ws.cell(row=rowNum,column=4,value="NA").font = Font(color='000000FF')
            if rowWrite[4] == None:
                ws.cell(row=rowNum,column=5,value="NA").font = Font(color='000000FF')  
            if rowWrite[5] == None:
                ws.cell(row=rowNum,column=6,value="NA").font = Font(color='000000FF')
            if rowWrite[6] == None:
                ws.cell(row=rowNum,column=7,value="NA").font = Font(color='000000FF')
            if rowWrite[7] == None:
                ws.cell(row=rowNum,column=8,value="NA").font = Font(color='000000FF')     

    for row in range(2,len(ws['A'])+1):
        for column in range(1,len(ExcelKeys)+1):
            ws.cell(row,column).border = thin_border   
            ws.cell(row,column).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True) 
            ws.row_dimensions[row].height = 15       

    refStr = "A1:I" + str(len(ws['A']))  
    tab = Table(displayName="Table1", ref=refStr)
    style = TableStyleInfo(name="TableStyleMedium6", showFirstColumn=False,
                        showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style                       
    ws.add_table(tab)

    for row in range(1,len(ws['A'])+1):
        for column in range(1,len(ExcelKeys)+1):
            if row == 1:
                ws.cell(row,column).fill = PatternFill(start_color='00008080',end_color='00008080', fill_type='solid')
            else:
                if row%2:
                    ws.cell(row,column).fill = PatternFill(start_color='00FFFFFF',end_color='00FFFFFF', fill_type='solid')
                else:
                    ws.cell(row,column).fill = PatternFill('solid', openpyxl.styles.colors.Color(theme=8, tint=0.8))

    return ws

# Compare Between Same View 
def compareSameView(wb,ws,IP_TYPE,FileJson,FileType):  
    """
    This function will do same veiw comparison.

    Usage:
        >>> import pinCompare
        >>> pinCompare.compareSameView("wb","ws","IP_TYPE","FileJson","FileType")

    :param: wb: workbook
    :param: ws: worksheet
    :param: IP_TYPE: ip type. ex, memory,io,logic etc.
    :param: FileJson: Lef, Lib or verilog files json dictionary
    :param: FileType: file type for same view comparison. Ex, lef,lib or verilog
    """
    FileNameKeys = [*FileJson.keys()]
    prevFileList = []
    if len(FileNameKeys) > 1:
        filePairs = getPairs(FileNameKeys)
        for pairs in filePairs:
            cellKeys_0 = [*FileJson[pairs[0]].keys()]
            cellKeys_1 = [*FileJson[pairs[1]].keys()]
            MergeCellKeys = sorted(list(set(cellKeys_0)|set(cellKeys_1)))
            for CellKey in MergeCellKeys:
                if CellKey in cellKeys_0 and CellKey in cellKeys_1:
                    PortKeys_0 = [*FileJson[pairs[0]][CellKey].keys()]
                    PortKeys_1 = [*FileJson[pairs[1]][CellKey].keys()]
                    MergePortKeys = sorted(list(set(PortKeys_0)|set(PortKeys_1)))    
                    for PortKey in MergePortKeys:
                        if PortKey in PortKeys_0 and PortKey in PortKeys_1:
                            PortDir_0 = FileJson[pairs[0]][CellKey][PortKey]["PortDirection"].upper()
                            PortDir_1 = FileJson[pairs[1]][CellKey][PortKey]["PortDirection"].upper()
                            if PortDir_0 == PortDir_1:
                                pass
                            else:
                                for rowNum, rowWrite in enumerate(ws.iter_rows(values_only=True), start=1):   
                                    if rowWrite[0] == CellKey and rowWrite[1] == PortKey: 
                                        # ERROR COMMENT
                                        if ws.cell(rowNum,len(ws[1])).value != None:
                                            ExistComment = "\n" + ws.cell(rowNum,len(ws[1])).value
                                        else :
                                            ExistComment = ""
                                        ws.cell(row=rowNum,column=len(ws[1]),value=str(f"File : {pairs[0]} & {pairs[1]}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction mismatched{ExistComment}")).font = Font(color='FF000000')
                                        #  
                                        break                                          

                        else:
                            if PortKey in PortKeys_0:
                                for rowNum, rowWrite in enumerate(ws.iter_rows(values_only=True), start=1):   
                                    if rowWrite[0] == CellKey and rowWrite[1] == PortKey: 
                                        # ERROR COMMENT
                                        if ws.cell(rowNum,len(ws[1])).value != None:
                                            ExistComment = "\n" + ws.cell(rowNum,len(ws[1])).value
                                        else :
                                            ExistComment = ""
                                        ws.cell(row=rowNum,column=len(ws[1]),value=str(f"File : {pairs[0]} & {pairs[1]}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' defined in {pairs[0]} but not in {pairs[1]}{ExistComment}")).font = Font(color='FF000000')
                                        #  
                                        break
                            if PortKey in PortKeys_1:
                                for rowNum, rowWrite in enumerate(ws.iter_rows(values_only=True), start=1):   
                                    if rowWrite[0] == CellKey and rowWrite[1] == PortKey: 
                                        # ERROR COMMENT
                                        if ws.cell(rowNum,len(ws[1])).value != None:
                                            ExistComment = "\n" + ws.cell(rowNum,len(ws[1])).value
                                        else :
                                            ExistComment = ""
                                        ws.cell(row=rowNum,column=len(ws[1]),value=str(f"File : {pairs[0]} & {pairs[1]}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' defined in {pairs[1]} but not in {pairs[0]}{ExistComment}")).font = Font(color='FF000000')
                                        # 
                                        break

                else:
                    if IP_TYPE != "MEMORY":
                        if CellKey in cellKeys_0:
                            for rowNum, rowWrite in enumerate(ws.iter_rows(values_only=True), start=1):   
                                if rowWrite[0] == CellKey: 
                                    # ERROR COMMENT
                                    if ws.cell(rowNum,len(ws[1])).value != None:
                                        ExistComment = "\n" + ws.cell(rowNum,len(ws[1])).value
                                    else :
                                        ExistComment = ""
                                    ws.cell(row=rowNum,column=len(ws[1]),value=str(f"File : {pairs[0]} & {pairs[1]}\nERROR : Cell '{CellKey}' defined in {pairs[0]} but not in {pairs[1]}{ExistComment}")).font = Font(color='FF000000')
                                    #                             
                        if CellKey in cellKeys_1:
                            for rowNum, rowWrite in enumerate(ws.iter_rows(values_only=True), start=1):   
                                if rowWrite[0] == CellKey: 
                                    # ERROR COMMENT
                                    if ws.cell(rowNum,len(ws[1])).value != None:
                                        ExistComment = "\n" + ws.cell(rowNum,len(ws[1])).value
                                    else :
                                        ExistComment = ""
                                    ws.cell(row=rowNum,column=len(ws[1]),value=str(f"File : {pairs[0]} & {pairs[1]}\nERROR : Cell '{CellKey}' defined in {pairs[1]} but not in {pairs[0]}{ExistComment}")).font = Font(color='FF000000')
                                    #  
                    else:
                        pass
    return wb,ws

# Get pair For same View Comparison
def getPairs(listFile):
    """
    This function will return file pair list for same view file comparison. 
    """
    out = []
    for i in range(len(listFile)-1):
        a = listFile.pop(0)
        for j in listFile:
            out.append([a, j])
    return out

# PIN COMPARE AND REPORT EXCEL
def PinCompareMain(LefJsonPath, LibJsonPath, VlogJsonPath, IP_TYPE, LefFileJsonPath, LibFileJsonPath, VlogFileJsonPath):
    # Logging.message("INFO", f"READING LEF JSON FROM FILE\n    {LefJsonPath}")
    try :
        lefJson = jsonRead(LefJsonPath)
    except:
        lefJson = ""
        Logging.message("WARNING", f"COULDN'T READ THE LEF JSON FROM FILE\n    {LefJsonPath}")

    # Logging.message("INFO", f"READING LIB JSON FROM FILE\n    {LibJsonPath}")
    try :
        libJson = jsonRead(LibJsonPath)
    except :
        libJson = ""
        Logging.message("WARNING", f"COULDN'T READ THE LIB JSON FROM FILE\n    {LibJsonPath}")

    # Logging.message("INFO", f"READING VLOG JSON FROM FILE\n    {VlogJsonPath}")        
    try :
        vlogJson = jsonRead(VlogJsonPath)
    except :
        vlogJson = ""
        Logging.message("WARNING", f"COULDN'T READ THE VLOG JSON FROM FILE\n    {VlogJsonPath}")

    # Logging.message("INFO", f"READING LEF File JSON FROM FILE\n    {LefFileJsonPath}")
    try :
        lefFileJson = jsonRead(LefFileJsonPath)
    except:
        lefFileJson = ""
        Logging.message("WARNING", f"COULDN'T READ THE LEF File JSON FROM FILE\n    {LefFileJsonPath}")        

    # Logging.message("INFO", f"READING LIB File JSON FROM FILE\n    {LibFileJsonPath}")
    try :
        libFileJson = jsonRead(LibFileJsonPath)
    except:
        libFileJson = ""
        Logging.message("WARNING", f"COULDN'T READ THE LIB File JSON FROM FILE\n    {LibFileJsonPath}") 

    # Logging.message("INFO", f"READING VLOG File JSON FROM FILE\n    {VlogFileJsonPath}")
    try :
        vlogFileJson = jsonRead(VlogFileJsonPath)
    except:
        vlogFileJson = ""
        Logging.message("WARNING", f"COULDN'T READ THE VLOG File JSON FROM FILE\n    {VlogFileJsonPath}") 

    wb,ws = compare(lefJson,libJson,vlogJson,IP_TYPE)

    if lefJson and lefFileJson:
        wb,ws = compareSameView(wb,ws,IP_TYPE, lefFileJson, "LEF")
    
    if libJson and libFileJson:
        wb,ws = compareSameView(wb,ws,IP_TYPE, libFileJson, "LIB") 

    if vlogJson and vlogFileJson:
        wb,ws = compareSameView(wb,ws,IP_TYPE, vlogFileJson, "VLOG")
    
    return wb