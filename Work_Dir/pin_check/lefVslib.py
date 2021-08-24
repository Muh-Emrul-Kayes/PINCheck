import os
import json
import sys
import re
from log import Logging
from utils import regexExtraction, write, jsonWrite,check_file, read,jsonRead, makeDirs,check_file
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment,PatternFill,Border,Side,Font 
from openpyxl.worksheet.table import Table, TableStyleInfo

# LEF Vs LIB Compare 
def compareLefVsLib(wb,ws,lefFileKeys, libFileKeys, lefJson, libJson, IP_TYPE, ExcelKeys):
    """
    This function will generate the lef vs lib comparison report as spreadsheet.

    Usage::
        >>> import lefVslib
        >>> lefVslib.compareLefVsLib("WorkBook","WorkSheet","lef_file_list","lib_file_list",
                                        "lef_json_data", "lib_json_data", "IP_TYPE", "Excel_heading")

    :param: wb: workbook 
    :param: ws: worksheet
    :param: lefFileKeys: lef file list 
    :param: libFileKeys: lib file list
    :param: lefJson: lef data as json dictionary format
    :param: libJson: lib data as json dictionary format
    :param: IP_TYPE: ip type. Ex. memory, io, logic etc.
    :param: ExcelKeys: spreadsheet heading as list 
    :returns: wb: workbook
    :returns: ws: worksheet
    """  
    Logging.message("INFO", "GENERATING LEF Vs LIB COMPARISON REPORT")
    for lefFile in lefFileKeys:
        for libFile in libFileKeys:
            lefCellKeys = [*lefJson[lefFile].keys()]
            libCellKeys = [*libJson[libFile].keys()]
            MergeCellKeys = sorted(list(set(lefCellKeys)|set(libCellKeys)))
            for CellKey in MergeCellKeys:
                if CellKey in lefCellKeys and CellKey in libCellKeys:
                    lefPortKeys = [*lefJson[lefFile][CellKey].keys()]
                    libPortKeys = [*libJson[libFile][CellKey].keys()]
                    MergePortKeys = sorted(list(set(lefPortKeys)|set(libPortKeys)))
                    # IF PIN NOT FOUND INSIDE LEF AND LIB CELL > COMMENT
                    if len(MergePortKeys) == 0:
                        FoundExistKey = False
                        for rowNum, rowWrite in enumerate(ws.iter_rows(values_only=True), start=1):
                            if rowWrite[0] == CellKey and rowWrite[1] == "NA":
                                # LEF DIR
                                ws.cell(row=rowNum,column=3,value="NA").font = Font(color='000000FF')
                                # LIB DIR
                                ws.cell(row=rowNum,column=4,value="NA").font = Font(color='000000FF')
                                # ERROR COMMENT
                                if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                    ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                else :
                                    ExistComment = ""
                                Commentlef = f"ERROR : For Cell '{CellKey}' Pin not defined in lef"
                                commentlib = f"ERROR : For Cell '{CellKey}' Pin not defined in lib"
                                ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {libFile}\n{Commentlef}\n{commentlib}")).font = Font(color='FF000000')                                              
                                # STATUS
                                ws.cell(row=rowNum,column=6,value="NA").font = Font(color='000000FF')
                                FoundExistKey = True
                                break                        
                        if not FoundExistKey:
                            FoundExistCell = False
                            for rowNum, rowWrite in enumerate(ws.iter_rows(values_only=True), start=1):
                                if rowWrite[0] == CellKey:
                                    ws.insert_rows(rowNum)
                                    ws.cell(row=rowNum,column=1,value=str(CellKey)).font = Font(color='FF000000')
                                    ws.cell(row=rowNum,column=2,value="NA").font = Font(color='000000FF')
                                    # LEF DIR
                                    ws.cell(row=rowNum,column=3,value="NA").font = Font(color='000000FF')
                                    # LIB DIR
                                    ws.cell(row=rowNum,column=4,value="NA").font = Font(color='000000FF')
                                    # ERROR COMMENT
                                    if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                        ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                    else :
                                        ExistComment = ""
                                    Commentlef = f"ERROR : For Cell '{CellKey}' Pin not defined in lef"
                                    commentlib = f"ERROR : For Cell '{CellKey}' Pin not defined in lib"
                                    ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {libFile}\n{Commentlef}\n{commentlib}")).font = Font(color='FF000000')                                                 
                                    ws.cell(row=rowNum,column=6,value="NA").font = Font(color='000000FF')
                                    FoundExistCell = True
                                    break 
                            if not FoundExistCell:
                                rowNum = len(ws['A'])+1
                                ws.cell(row=rowNum,column=1,value=str(CellKey)).font = Font(color='FF000000')
                                ws.cell(row=rowNum,column=2,value="NA").font = Font(color='000000FF')
                                # LEF DIR
                                ws.cell(row=rowNum,column=3,value="NA").font = Font(color='000000FF')
                                # LIB DIR
                                ws.cell(row=rowNum,column=4,value="NA").font = Font(color='000000FF')
                                # ERROR COMMENT
                                if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                    ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                else :
                                    ExistComment = ""
                                Commentlef = f"ERROR : For Cell '{CellKey}' Pin not defined in lef"
                                commentlib = f"ERROR : For Cell '{CellKey}' Pin not defined in lib"
                                ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {libFile}\n{Commentlef}\n{commentlib}")).font = Font(color='FF000000')                                                
                                ws.cell(row=rowNum,column=6,value="NA").font = Font(color='000000FF')
                    # Compare PIN Between LEF and LIB CELL            
                    for PortKey in MergePortKeys:
                        if PortKey in lefPortKeys and PortKey in libPortKeys:
                            lefPortDir = lefJson[lefFile][CellKey][PortKey]["PortDirection"].upper()
                            libPortDir = libJson[libFile][CellKey][PortKey]["PortDirection"].upper()
                            if lefPortDir == libPortDir and lefPortDir.strip() and libPortDir.strip():               
                                FoundExistKey = False
                                for rowNum, rowWrite in enumerate(ws.iter_rows(values_only=True), start=1):
                                    if rowWrite[0] == CellKey and rowWrite[1] == PortKey:
                                        # LEF DIR
                                        if (rowWrite[2] == None or rowWrite[2] == "NA") and rowWrite[2] != "ERROR":
                                            if lefPortDir.strip():
                                                ws.cell(row=rowNum,column=3,value=str(lefPortDir.upper())).font = Font(color='FF000000')
                                            else:
                                                ws.cell(row=rowNum,column=3,value="ERROR").font = Font(color='00FF0000')
                                                # ERROR COMMENT
                                                if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                    ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                                else :
                                                    ExistComment = ""
                                                ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {libFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in lef")).font = Font(color='FF000000')
                                        else : 
                                            if rowWrite[2] == lefPortDir.upper():
                                                pass
                                            else :
                                                ws.cell(row=rowNum,column=3,value="ERROR").font = Font(color='00FF0000')
                                                # ERROR COMMENT
                                                if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                    ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                                else :
                                                    ExistComment = ""
                                                if lefPortDir.strip():
                                                    ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {libFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction mismatched within lef")).font = Font(color='FF000000')
                                                else:
                                                    ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {libFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in lef")).font = Font(color='FF000000')
                                        # LIB DIR
                                        if (rowWrite[3] == None or rowWrite[3] == "NA") and rowWrite[3] != "ERROR":
                                            if libPortDir.strip():
                                                ws.cell(row=rowNum,column=4,value=str(libPortDir.upper())).font = Font(color='FF000000')
                                            else:
                                                ws.cell(row=rowNum,column=4,value="ERROR").font = Font(color='00FF0000')
                                                # ERROR COMMENT
                                                if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                    ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                                else :
                                                    ExistComment = ""
                                                ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {libFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in lib")).font = Font(color='FF000000')                                               
                                        else : 
                                            if rowWrite[3] == libPortDir.upper():
                                                pass
                                            else :
                                                ws.cell(row=rowNum,column=4,value="ERROR").font = Font(color='00FF0000') 
                                                # ERROR COMMENT
                                                if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                    ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                                else :
                                                    ExistComment = ""
                                                if libPortDir.strip():
                                                    ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {libFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction mismatched within lib")).font = Font(color='FF000000')
                                                else:
                                                    ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {libFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in lib")).font = Font(color='FF000000')                                                
                                        # STATUS
                                        if rowWrite[5] == None or rowWrite[5] == "NA":
                                            ws.cell(row=rowNum,column=6,value="PASS").font = Font(color='00008000')
                                        FoundExistKey = True
                                        break                        
                                if not FoundExistKey:
                                    FoundExistCell = False
                                    for rowNum, rowWrite in enumerate(ws.iter_rows(values_only=True), start=1):
                                        if rowWrite[0] == CellKey:
                                            ws.insert_rows(rowNum)
                                            ws.cell(row=rowNum,column=1,value=str(CellKey)).font = Font(color='FF000000')
                                            ws.cell(row=rowNum,column=2,value=str(PortKey)).font = Font(color='FF000000')
                                            if lefPortDir.strip():
                                                ws.cell(row=rowNum,column=3,value=str(lefPortDir.upper())).font = Font(color='FF000000')
                                            else:
                                                ws.cell(row=rowNum,column=3,value="ERROR").font = Font(color='00FF0000')
                                                # ERROR COMMENT
                                                if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                    ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                                else :
                                                    ExistComment = ""
                                                ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {libFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in lef")).font = Font(color='FF000000')
                                                #                                               
                                            if libPortDir.strip():
                                                ws.cell(row=rowNum,column=4,value=str(libPortDir.upper())).font = Font(color='FF000000')
                                            else:
                                                ws.cell(row=rowNum,column=4,value="ERROR").font = Font(color='00FF0000')
                                                # ERROR COMMENT
                                                if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                    ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                                else :
                                                    ExistComment = ""
                                                ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {libFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in lib")).font = Font(color='FF000000')
                                                #                                                 
                                            ws.cell(row=rowNum,column=6,value="PASS").font = Font(color='00008000')
                                            FoundExistCell = True
                                            break 
                                    if not FoundExistCell:
                                        rowNum = len(ws['A'])+1
                                        ws.cell(row=rowNum,column=1,value=str(CellKey)).font = Font(color='FF000000')
                                        ws.cell(row=rowNum,column=2,value=str(PortKey)).font = Font(color='FF000000')
                                        if lefPortDir.strip():
                                            ws.cell(row=rowNum,column=3,value=str(lefPortDir.upper())).font = Font(color='FF000000')
                                        else:
                                            ws.cell(row=rowNum,column=3,value="ERROR").font = Font(color='00FF0000')
                                            # ERROR COMMENT
                                            if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                            else :
                                                ExistComment = ""
                                            ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {libFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in lef")).font = Font(color='FF000000')
                                            #                                             
                                        if libPortDir.strip():
                                            ws.cell(row=rowNum,column=4,value=str(libPortDir.upper())).font = Font(color='FF000000')
                                        else:
                                            ws.cell(row=rowNum,column=4,value="ERROR").font = Font(color='00FF0000')
                                            # ERROR COMMENT
                                            if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                            else :
                                                ExistComment = ""
                                            ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {libFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in lib")).font = Font(color='FF000000')
                                            #                                             
                                        ws.cell(row=rowNum,column=6,value="PASS").font = Font(color='00008000')
                                #
                                                        
                            else:
                                #
                                FoundExistKey = False
                                for rowNum, rowWrite in enumerate(ws.iter_rows(values_only=True), start=1):
                                    if rowWrite[0] == CellKey and rowWrite[1] == PortKey:
                                        # LEF DIR
                                        if (rowWrite[2] == None or rowWrite[2] == "NA") and rowWrite[2] != "ERROR":
                                            if lefPortDir.strip():
                                                ws.cell(row=rowNum,column=3,value=str(lefPortDir.upper())).font = Font(color='FF000000')
                                            else:
                                                ws.cell(row=rowNum,column=3,value="ERROR").font = Font(color='00FF0000')
                                                # ERROR COMMENT
                                                if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                    ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                                else :
                                                    ExistComment = ""
                                                ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {libFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in lef")).font = Font(color='FF000000')
                                                #                                                 
                                        else : 
                                            if rowWrite[2] == lefPortDir.upper():
                                                pass
                                            else :
                                                ws.cell(row=rowNum,column=3,value="ERROR").font = Font(color='00FF0000')
                                                # ERROR COMMENT
                                                if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                    ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                                else :
                                                    ExistComment = ""
                                                if lefPortDir.strip():
                                                    ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {libFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction mismatched within lef")).font = Font(color='FF000000')
                                                else:
                                                    ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {libFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in lef")).font = Font(color='FF000000')
                                                #                                                
                                        # LIB DIR
                                        if (rowWrite[3] == None or rowWrite[3] == "NA") and rowWrite[3] != "ERROR":
                                            if libPortDir.strip():
                                                ws.cell(row=rowNum,column=4,value=str(libPortDir.upper())).font = Font(color='FF000000')
                                            else:
                                                ws.cell(row=rowNum,column=4,value="ERROR").font = Font(color='00FF0000')
                                                # ERROR COMMENT
                                                if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                    ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                                else :
                                                    ExistComment = ""
                                                ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {libFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in lib")).font = Font(color='FF000000')
                                                #                                                 
                                        else : 
                                            if rowWrite[3] == libPortDir.upper():
                                                pass
                                            else :
                                                ws.cell(row=rowNum,column=4,value="ERROR").font = Font(color='00FF0000')
                                                # ERROR COMMENT
                                                if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                    ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                                else :
                                                    ExistComment = ""
                                                if libPortDir.strip():
                                                    ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {libFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction mismatched within lib")).font = Font(color='FF000000')
                                                else:
                                                    ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {libFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in lib")).font = Font(color='FF000000')
                                                #                                                 
                                        # STATUS
                                        if rowWrite[5] == None or rowWrite[5] == "NA" or rowWrite[5] == "PASS":
                                            ws.cell(row=rowNum,column=6,value="ERROR").font = Font(color='00FF0000')
                                        # COMMENT
                                        if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                            ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                        else :
                                            ExistComment = ""
                                        ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {libFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction mismatched between lef and lib")).font = Font(color='FF000000')                                                                                        
                                        FoundExistKey = True
                                        break                        
                                if not FoundExistKey:
                                    FoundExistCell = False
                                    for rowNum, rowWrite in enumerate(ws.iter_rows(values_only=True), start=1):
                                        if rowWrite[0] == CellKey:
                                            ws.insert_rows(rowNum)
                                            ws.cell(row=rowNum,column=1,value=str(CellKey)).font = Font(color='FF000000')
                                            ws.cell(row=rowNum,column=2,value=str(PortKey)).font = Font(color='FF000000')
                                            if lefPortDir.strip():
                                                ws.cell(row=rowNum,column=3,value=str(lefPortDir.upper())).font = Font(color='FF000000')
                                            else:
                                                ws.cell(row=rowNum,column=3,value="ERROR").font = Font(color='00FF0000')
                                                # ERROR COMMENT
                                                if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                    ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                                else :
                                                    ExistComment = ""
                                                ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {libFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in lef")).font = Font(color='FF000000')
                                                #                                                 
                                            if libPortDir.strip():
                                                ws.cell(row=rowNum,column=4,value=str(libPortDir.upper())).font = Font(color='FF000000')
                                            else:
                                                ws.cell(row=rowNum,column=4,value="ERROR").font = Font(color='00FF0000')
                                                # ERROR COMMENT
                                                if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                    ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                                else :
                                                    ExistComment = ""
                                                ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {libFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in lib")).font = Font(color='FF000000')
                                                #                                                 
                                            ws.cell(row=rowNum,column=6,value="ERROR").font = Font(color='00FF0000')
                                            # COMMENT
                                            if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                            else :
                                                ExistComment = ""
                                            ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {libFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction mismatched between lef and lib")).font = Font(color='FF000000')                                            
                                            FoundExistCell = True
                                            break 
                                    if not FoundExistCell:
                                        rowNum = len(ws['A'])+1
                                        ws.cell(row=rowNum,column=1,value=str(CellKey)).font = Font(color='FF000000')
                                        ws.cell(row=rowNum,column=2,value=str(PortKey)).font = Font(color='FF000000')
                                        if lefPortDir.strip():
                                            ws.cell(row=rowNum,column=3,value=str(lefPortDir.upper())).font = Font(color='FF000000')
                                        else:
                                            ws.cell(row=rowNum,column=3,value="ERROR").font = Font(color='00FF0000')
                                            # ERROR COMMENT
                                            if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                            else :
                                                ExistComment = ""
                                            ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {libFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in lef")).font = Font(color='FF000000')
                                            #                                             
                                        if libPortDir.strip():
                                            ws.cell(row=rowNum,column=4,value=str(libPortDir.upper())).font = Font(color='FF000000')
                                        else:
                                            ws.cell(row=rowNum,column=4,value="ERROR").font = Font(color='00FF0000')
                                            # ERROR COMMENT
                                            if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                            else :
                                                ExistComment = ""
                                            ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {libFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in lib")).font = Font(color='FF000000')
                                            #                                             
                                        ws.cell(row=rowNum,column=6,value="ERROR").font = Font(color='00FF0000')
                                        # COMMENT
                                        if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                            ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                        else :
                                            ExistComment = ""
                                        ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {libFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction mismatched between lef and lib")).font = Font(color='FF000000')                                         
                                #                                
                        # FOR PORT NOT FOUND IN CELL BETWEEN LEF AND LIB
                        else:
                            if PortKey in lefPortKeys:
                                lefPortDir = lefJson[lefFile][CellKey][PortKey]["PortDirection"].upper()  
                                #
                                FoundExistKey = False
                                for rowNum, rowWrite in enumerate(ws.iter_rows(values_only=True), start=1):
                                    if rowWrite[0] == CellKey and rowWrite[1] == PortKey:
                                        # LEF DIR
                                        if (rowWrite[2] == None or rowWrite[2] == "NA") and rowWrite[2] != "ERROR":
                                            if lefPortDir.strip():
                                                ws.cell(row=rowNum,column=3,value=str(lefPortDir.upper())).font = Font(color='FF000000')
                                            else:
                                                ws.cell(row=rowNum,column=3,value="ERROR").font = Font(color='00FF0000')
                                                # ERROR COMMENT
                                                if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                    ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                                else :
                                                    ExistComment = ""
                                                ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {libFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in lef")).font = Font(color='FF000000')
                                                #                                                 
                                        else : 
                                            if rowWrite[2] == lefPortDir.upper():
                                                pass
                                            else :
                                                ws.cell(row=rowNum,column=3,value="ERROR").font = Font(color='00FF0000')
                                                # ERROR COMMENT
                                                if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                    ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                                else :
                                                    ExistComment = ""
                                                if lefPortDir.strip():
                                                    ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {libFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction mismatched within lef")).font = Font(color='FF000000')
                                                else:
                                                    ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {libFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in lef")).font = Font(color='FF000000')
                                                #                                                
                                        # LIB DIR
                                        if rowWrite[3] == None:
                                            ws.cell(row=rowNum,column=4,value="NA").font = Font(color='000000FF') 
                                        # STATUS
                                        # if rowWrite[5] != "PASS":
                                        #     ws.cell(row=rowNum,column=6,value="FAIL").font = Font(color='00FF0000')
                                        
                                        ws.cell(row=rowNum,column=6,value="FAIL").font = Font(color='00FF0000')                                          
                                        # COMMENT
                                        if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                            ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                        else :
                                            ExistComment = ""
                                        ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {libFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' defined in lef but not in lib")).font = Font(color='FF000000')
                                        FoundExistKey = True
                                        break                        
                                if not FoundExistKey:
                                    FoundExistCell = False
                                    for rowNum, rowWrite in enumerate(ws.iter_rows(values_only=True), start=1):
                                        if rowWrite[0] == CellKey:
                                            ws.insert_rows(rowNum)
                                            ws.cell(row=rowNum,column=1,value=str(CellKey)).font = Font(color='FF000000')
                                            ws.cell(row=rowNum,column=2,value=str(PortKey)).font = Font(color='FF000000')
                                            if lefPortDir.strip():
                                                ws.cell(row=rowNum,column=3,value=str(lefPortDir.upper())).font = Font(color='FF000000')
                                            else:
                                                ws.cell(row=rowNum,column=3,value="ERROR").font = Font(color='00FF0000')
                                                # ERROR COMMENT
                                                if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                    ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                                else :
                                                    ExistComment = ""
                                                ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {libFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in lef")).font = Font(color='FF000000')
                                                #                                                 
                                            ws.cell(row=rowNum,column=4,value="NA").font = Font(color='000000FF')
                                            ws.cell(row=rowNum,column=6,value="FAIL").font = Font(color='00FF0000')
                                            # COMMENT
                                            if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                            else :
                                                ExistComment = ""
                                            ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {libFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' defined in lef but not in lib")).font = Font(color='FF000000')                                            
                                            FoundExistCell = True
                                            break 
                                    if not FoundExistCell:
                                        rowNum = len(ws['A'])+1
                                        ws.cell(row=rowNum,column=1,value=str(CellKey)).font = Font(color='FF000000')
                                        ws.cell(row=rowNum,column=2,value=str(PortKey)).font = Font(color='FF000000')
                                        if lefPortDir.strip():
                                            ws.cell(row=rowNum,column=3,value=str(lefPortDir.upper())).font = Font(color='FF000000')
                                        else:
                                            ws.cell(row=rowNum,column=3,value="ERROR").font = Font(color='00FF0000')
                                            # ERROR COMMENT
                                            if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                            else :
                                                ExistComment = ""
                                            ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {libFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in lef")).font = Font(color='FF000000')
                                            #                                             
                                        ws.cell(row=rowNum,column=4,value="NA").font = Font(color='000000FF')
                                        ws.cell(row=rowNum,column=6,value="FAIL").font = Font(color='00FF0000')
                                        # COMMENT
                                        if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                            ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                        else :
                                            ExistComment = ""
                                        ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {libFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' defined in lef but not in lib")).font = Font(color='FF000000')                                            
                                #                                   
                            else: 
                                libPortDir = libJson[libFile][CellKey][PortKey]["PortDirection"].upper()   
                                #
                                FoundExistKey = False
                                for rowNum, rowWrite in enumerate(ws.iter_rows(values_only=True), start=1):
                                    if rowWrite[0] == CellKey and rowWrite[1] == PortKey:
                                        # LEF DIR
                                        if rowWrite[2] == None:
                                            ws.cell(row=rowNum,column=3,value="NA").font = Font(color='000000FF')
                                        # LIB DIR 
                                        if (rowWrite[3] == None or rowWrite[3] == "NA") and rowWrite[3] != "ERROR":
                                            if libPortDir.strip():
                                                ws.cell(row=rowNum,column=4,value=str(libPortDir.upper())).font = Font(color='FF000000')
                                            else:
                                                ws.cell(row=rowNum,column=4,value="ERROR").font = Font(color='00FF0000')
                                                # ERROR COMMENT
                                                if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                    ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                                else :
                                                    ExistComment = ""
                                                ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {libFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in lib")).font = Font(color='FF000000')
                                                #                                                 
                                        else : 
                                            if rowWrite[3] == libPortDir.upper():
                                                pass
                                            else :
                                                ws.cell(row=rowNum,column=4,value="ERROR").font = Font(color='00FF0000')         
                                                # ERROR COMMENT
                                                if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                    ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                                else :
                                                    ExistComment = ""
                                                if libPortDir.strip():
                                                    ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {libFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction mismatched within lib")).font = Font(color='FF000000')
                                                else:
                                                    ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {libFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in lib")).font = Font(color='FF000000')
                                                #                                                                                   
                                        # STATUS
                                        # if rowWrite[5] != "PASS":
                                        #     ws.cell(row=rowNum,column=6,value="FAIL").font = Font(color='00FF0000')
                                        
                                        ws.cell(row=rowNum,column=6,value="FAIL").font = Font(color='00FF0000')
                                        # COMMENT
                                        if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                            ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                        else :
                                            ExistComment = ""
                                        ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {libFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' defined in lib but not in lef")).font = Font(color='FF000000')
                                        FoundExistKey = True
                                        break                        
                                if not FoundExistKey:
                                    FoundExistCell = False
                                    for rowNum, rowWrite in enumerate(ws.iter_rows(values_only=True), start=1):
                                        if rowWrite[0] == CellKey:
                                            ws.insert_rows(rowNum)
                                            ws.cell(row=rowNum,column=1,value=str(CellKey)).font = Font(color='FF000000')
                                            ws.cell(row=rowNum,column=2,value=str(PortKey)).font = Font(color='FF000000')
                                            ws.cell(row=rowNum,column=3,value="NA").font = Font(color='000000FF')
                                            if libPortDir.strip():
                                                ws.cell(row=rowNum,column=4,value=str(libPortDir.upper())).font = Font(color='FF000000')
                                            else:
                                                ws.cell(row=rowNum,column=4,value="ERROR").font = Font(color='00FF0000')
                                                # ERROR COMMENT
                                                if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                    ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                                else :
                                                    ExistComment = ""
                                                ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {libFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in lib")).font = Font(color='FF000000')
                                                #                                                 
                                            ws.cell(row=rowNum,column=6,value="FAIL").font = Font(color='00FF0000')
                                            # COMMENT
                                            if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                            else :
                                                ExistComment = ""
                                            ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {libFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' defined in lib but not in lef")).font = Font(color='FF000000')                                            
                                            FoundExistCell = True
                                            break 
                                    if not FoundExistCell:
                                        rowNum = len(ws['A'])+1
                                        ws.cell(row=rowNum,column=1,value=str(CellKey)).font = Font(color='FF000000')
                                        ws.cell(row=rowNum,column=2,value=str(PortKey)).font = Font(color='FF000000')
                                        ws.cell(row=rowNum,column=3,value="NA").font = Font(color='000000FF')
                                        if libPortDir.strip():
                                            ws.cell(row=rowNum,column=4,value=str(libPortDir.upper())).font = Font(color='FF000000')
                                        else:
                                            ws.cell(row=rowNum,column=4,value="ERROR").font = Font(color='00FF0000')
                                            # ERROR COMMENT
                                            if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                            else :
                                                ExistComment = ""
                                            ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {libFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in lib")).font = Font(color='FF000000')
                                            #                                             
                                        ws.cell(row=rowNum,column=6,value="FAIL").font = Font(color='00FF0000')
                                        # COMMENT
                                        if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                            ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                        else :
                                            ExistComment = ""
                                        ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {libFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' defined in lib but not in lef")).font = Font(color='FF000000')                                            
                                #                                       
                # FOR CELL NOT FOUND IN ONE FILE LEF OR LIB        
                else:
                    if CellKey in lefCellKeys:
                        lefPortKeys = [*lefJson[lefFile][CellKey].keys()]
                        # IF PORT NOT FOUND IN LEF > ERROR COMMENT
                        if len(lefPortKeys) == 0:
                            #
                            FoundExistKey = False
                            for rowNum, rowWrite in enumerate(ws.iter_rows(values_only=True), start=1):
                                if rowWrite[0] == CellKey and rowWrite[1] == "NA":
                                    # lef DIR
                                    ws.cell(row=rowNum,column=3,value="NA").font = Font(color='000000FF')
                                    # lib DIR
                                    ws.cell(row=rowNum,column=4,value="NA").font = Font(color='000000FF')
                                    # ERROR COMMENT
                                    if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                        ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                    else :
                                        ExistComment = ""
                                    Commentlef = f"ERROR : For Cell '{CellKey}' Pin not defined in lef"
                                    commentlib = f"ERROR : Cell '{CellKey}' not defined in lib"
                                    ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {libFile}\n{Commentlef}\n{commentlib}")).font = Font(color='FF000000')
                                    #                                                
                                    # STATUS
                                    ws.cell(row=rowNum,column=6,value="FAIL").font = Font(color='00FF0000')
                                    FoundExistKey = True
                                    break                        
                            if not FoundExistKey:
                                FoundExistCell = False
                                for rowNum, rowWrite in enumerate(ws.iter_rows(values_only=True), start=1):
                                    if rowWrite[0] == CellKey:
                                        ws.insert_rows(rowNum)
                                        ws.cell(row=rowNum,column=1,value=str(CellKey)).font = Font(color='FF000000')
                                        ws.cell(row=rowNum,column=2,value="NA").font = Font(color='000000FF')
                                        # lef DIR
                                        ws.cell(row=rowNum,column=3,value="NA").font = Font(color='000000FF')
                                        # lib DIR
                                        ws.cell(row=rowNum,column=4,value="NA").font = Font(color='000000FF')
                                        # ERROR COMMENT
                                        if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                            ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                        else :
                                            ExistComment = ""
                                        Commentlef = f"ERROR : For Cell '{CellKey}' Pin not defined in lef"
                                        commentlib = f"ERROR : Cell '{CellKey}' not defined in lib"
                                        ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {libFile}\n{Commentlef}\n{commentlib}")).font = Font(color='FF000000')
                                        #                                                
                                        # STATUS
                                        ws.cell(row=rowNum,column=6,value="FAIL").font = Font(color='00FF0000')
                                        FoundExistCell = True
                                        break 
                                if not FoundExistCell:
                                    rowNum = len(ws['A'])+1
                                    ws.cell(row=rowNum,column=1,value=str(CellKey)).font = Font(color='FF000000')
                                    ws.cell(row=rowNum,column=2,value="NA").font = Font(color='000000FF')
                                    # lef DIR
                                    ws.cell(row=rowNum,column=3,value="NA").font = Font(color='000000FF')
                                    # lib DIR
                                    ws.cell(row=rowNum,column=4,value="NA").font = Font(color='000000FF')
                                    # ERROR COMMENT
                                    if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                        ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                    else :
                                        ExistComment = ""
                                    Commentlef = f"ERROR : For Cell '{CellKey}' Pin not defined in lef"
                                    commentlib = f"ERROR : Cell '{CellKey}' not defined in lib"
                                    ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {libFile}\n{Commentlef}\n{commentlib}")).font = Font(color='FF000000')
                                    #                                                
                                    # STATUS
                                    ws.cell(row=rowNum,column=6,value="FAIL").font = Font(color='00FF0000')
                            #
                        for PortKey in lefPortKeys:
                            lefPortDir = lefJson[lefFile][CellKey][PortKey]["PortDirection"].upper()   
                            #
                            FoundExistKey = False
                            for rowNum, rowWrite in enumerate(ws.iter_rows(values_only=True), start=1):
                                if rowWrite[0] == CellKey and rowWrite[1] == PortKey:
                                    # LEF DIR
                                    if (rowWrite[2] == None or rowWrite[2] == "NA") and rowWrite[2] != "ERROR":
                                        if lefPortDir.strip():
                                            ws.cell(row=rowNum,column=3,value=str(lefPortDir.upper())).font = Font(color='FF000000')
                                        else:
                                            ws.cell(row=rowNum,column=3,value="ERROR").font = Font(color='00FF0000')
                                            # ERROR COMMENT
                                            if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                            else :
                                                ExistComment = ""
                                            ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {libFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in lef")).font = Font(color='FF000000')
                                            #                                             
                                    else : 
                                        if rowWrite[2] == lefPortDir.upper():
                                            pass
                                        else :
                                            ws.cell(row=rowNum,column=3,value="ERROR").font = Font(color='00FF0000')  
                                            # ERROR COMMENT
                                            if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                            else :
                                                ExistComment = ""
                                            if lefPortDir.strip():
                                                ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {libFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction mismatched within lef")).font = Font(color='FF000000')
                                            else:
                                                ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {libFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in lef")).font = Font(color='FF000000')
                                            #                                                                               
                                    # LIB DIR 
                                    if rowWrite[3] == None:
                                        ws.cell(row=rowNum,column=4,value="NA").font = Font(color='000000FF')                                             
                                    # STATUS
                                    # if rowWrite[5] != "PASS":
                                    #     ws.cell(row=rowNum,column=6,value="FAIL").font = Font(color='00FF0000')
                                    
                                    ws.cell(row=rowNum,column=6,value="FAIL").font = Font(color='00FF0000')  
                                    # COMMENT
                                    if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                        ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                    else :
                                        ExistComment = ""
                                    ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {libFile}\nERROR : Cell '{CellKey}' defined in lef but not in lib")).font = Font(color='FF000000')
                                    FoundExistKey = True
                                    break                        
                            if not FoundExistKey:
                                FoundExistCell = False
                                for rowNum, rowWrite in enumerate(ws.iter_rows(values_only=True), start=1):
                                    if rowWrite[0] == CellKey:
                                        ws.insert_rows(rowNum)
                                        ws.cell(row=rowNum,column=1,value=str(CellKey)).font = Font(color='FF000000')
                                        ws.cell(row=rowNum,column=2,value=str(PortKey)).font = Font(color='FF000000')
                                        if lefPortDir.strip():
                                            ws.cell(row=rowNum,column=3,value=str(lefPortDir.upper())).font = Font(color='FF000000')
                                        else:
                                            ws.cell(row=rowNum,column=3,value="ERROR").font = Font(color='00FF0000')
                                            # ERROR COMMENT
                                            if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                            else :
                                                ExistComment = ""
                                            ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {libFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in lef")).font = Font(color='FF000000')
                                            #                                            
                                        ws.cell(row=rowNum,column=4,value="NA").font = Font(color='000000FF')
                                        ws.cell(row=rowNum,column=6,value="FAIL").font = Font(color='00FF0000')
                                        # COMMENT
                                        if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                            ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                        else :
                                            ExistComment = ""
                                        ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {libFile}\nERROR : Cell '{CellKey}' defined in lef but not in lib")).font = Font(color='FF000000')                                            
                                        FoundExistCell = True
                                        break 
                                if not FoundExistCell:
                                    rowNum = len(ws['A'])+1
                                    ws.cell(row=rowNum,column=1,value=str(CellKey)).font = Font(color='FF000000')
                                    ws.cell(row=rowNum,column=2,value=str(PortKey)).font = Font(color='FF000000')
                                    if lefPortDir.strip():
                                        ws.cell(row=rowNum,column=3,value=str(lefPortDir.upper())).font = Font(color='FF000000')
                                    else:
                                        ws.cell(row=rowNum,column=3,value="ERROR").font = Font(color='00FF0000')
                                        # ERROR COMMENT
                                        if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                            ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                        else :
                                            ExistComment = ""
                                        ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {libFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in lef")).font = Font(color='FF000000')
                                        #                                        
                                    ws.cell(row=rowNum,column=4,value="NA").font = Font(color='000000FF')
                                    ws.cell(row=rowNum,column=6,value="FAIL").font = Font(color='00FF0000')
                                    # COMMENT
                                    if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                        ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                    else :
                                        ExistComment = ""
                                    ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {libFile}\nERROR : Cell '{CellKey}' defined in lef but not in lib")).font = Font(color='FF000000')                                            
                            #                        
                    else: 
                        libPortKeys = [*libJson[libFile][CellKey].keys()]
                        # IF PORT NOT FOUND IN LIB > ERROR COMMENT
                        if len(libPortKeys) == 0:
                            #
                            FoundExistKey = False
                            for rowNum, rowWrite in enumerate(ws.iter_rows(values_only=True), start=1):
                                if rowWrite[0] == CellKey and rowWrite[1] == "NA":
                                    # lef DIR
                                    ws.cell(row=rowNum,column=3,value="NA").font = Font(color='000000FF')
                                    # lib DIR
                                    ws.cell(row=rowNum,column=4,value="NA").font = Font(color='000000FF')
                                    # ERROR COMMENT
                                    if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                        ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                    else :
                                        ExistComment = ""
                                    Commentlef = f"ERROR : Cell '{CellKey}' not defined in lef"
                                    commentlib = f"ERROR : For Cell '{CellKey}' Pin not defined in lib"
                                    ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {libFile}\n{Commentlef}\n{commentlib}")).font = Font(color='FF000000')
                                    #                                                
                                    # STATUS
                                    ws.cell(row=rowNum,column=6,value="FAIL").font = Font(color='00FF0000')
                                    FoundExistKey = True
                                    break                        
                            if not FoundExistKey:
                                FoundExistCell = False
                                for rowNum, rowWrite in enumerate(ws.iter_rows(values_only=True), start=1):
                                    if rowWrite[0] == CellKey:
                                        ws.insert_rows(rowNum)
                                        ws.cell(row=rowNum,column=1,value=str(CellKey)).font = Font(color='FF000000')
                                        ws.cell(row=rowNum,column=2,value="NA").font = Font(color='000000FF')
                                        # lef DIR
                                        ws.cell(row=rowNum,column=3,value="NA").font = Font(color='000000FF')
                                        # lib DIR
                                        ws.cell(row=rowNum,column=4,value="NA").font = Font(color='000000FF')
                                        # ERROR COMMENT
                                        if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                            ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                        else :
                                            ExistComment = ""
                                        Commentlef = f"ERROR : Cell '{CellKey}' not defined in lef"
                                        commentlib = f"ERROR : For Cell '{CellKey}' Pin not defined in lib"
                                        ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {libFile}\n{Commentlef}\n{commentlib}")).font = Font(color='FF000000')
                                        #                                                
                                        # STATUS
                                        ws.cell(row=rowNum,column=6,value="FAIL").font = Font(color='00FF0000') 
                                        FoundExistCell = True
                                        break 
                                if not FoundExistCell:
                                    rowNum = len(ws['A'])+1
                                    ws.cell(row=rowNum,column=1,value=str(CellKey)).font = Font(color='FF000000')
                                    ws.cell(row=rowNum,column=2,value="NA").font = Font(color='000000FF')
                                    # lef DIR
                                    ws.cell(row=rowNum,column=3,value="NA").font = Font(color='000000FF')
                                    # lib DIR
                                    ws.cell(row=rowNum,column=4,value="NA").font = Font(color='000000FF')
                                    # ERROR COMMENT
                                    if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                        ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                    else :
                                        ExistComment = ""
                                    Commentlef = f"ERROR : Cell '{CellKey}' not defined in lef"
                                    commentlib = f"ERROR : For Cell '{CellKey}' Pin not defined in lib"
                                    ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {libFile}\n{Commentlef}\n{commentlib}")).font = Font(color='FF000000')
                                    #                                                
                                    # STATUS
                                    ws.cell(row=rowNum,column=6,value="FAIL").font = Font(color='00FF0000')
                            #
                        for PortKey in libPortKeys:
                            libPortDir = libJson[libFile][CellKey][PortKey]["PortDirection"].upper()   
                            #
                            FoundExistKey = False
                            for rowNum, rowWrite in enumerate(ws.iter_rows(values_only=True), start=1):
                                if rowWrite[0] == CellKey and rowWrite[1] == PortKey:
                                    # LEF DIR
                                    if rowWrite[2] == None:
                                        ws.cell(row=rowNum,column=3,value="NA").font = Font(color='000000FF')                                   
                                    # LIB DIR 
                                    if (rowWrite[3] == None or rowWrite[3] == "NA") and rowWrite[3] != "ERROR":
                                        if libPortDir.strip():
                                            ws.cell(row=rowNum,column=4,value=str(libPortDir.upper())).font = Font(color='FF000000')
                                        else:
                                            ws.cell(row=rowNum,column=4,value="ERROR").font = Font(color='00FF0000')
                                            # ERROR COMMENT
                                            if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                            else :
                                                ExistComment = ""
                                            ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {libFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in lib")).font = Font(color='FF000000')
                                            #                                            
                                    else : 
                                        if rowWrite[3] == libPortDir.upper():
                                            pass
                                        else :
                                            ws.cell(row=rowNum,column=4,value="ERROR").font = Font(color='00FF0000')
                                            # ERROR COMMENT
                                            if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                            else :
                                                ExistComment = ""
                                            if libPortDir.strip():
                                                ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {libFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction mismatched within lib")).font = Font(color='FF000000')
                                            else:
                                                ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {libFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in lib")).font = Font(color='FF000000')
                                            #
                                    # STATUS
                                    # if rowWrite[5] != "PASS":
                                    #     ws.cell(row=rowNum,column=6,value="FAIL").font = Font(color='00FF0000')
                                    
                                    ws.cell(row=rowNum,column=6,value="FAIL").font = Font(color='00FF0000')  
                                    # COMMENT
                                    if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                        ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                    else :
                                        ExistComment = ""
                                    ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {libFile}\nERROR : Cell '{CellKey}' defined in lib but not in lef")).font = Font(color='FF000000')
                                    FoundExistKey = True
                                    break                        
                            if not FoundExistKey:
                                FoundExistCell = False
                                for rowNum, rowWrite in enumerate(ws.iter_rows(values_only=True), start=1):
                                    if rowWrite[0] == CellKey:
                                        ws.insert_rows(rowNum)
                                        ws.cell(row=rowNum,column=1,value=str(CellKey)).font = Font(color='FF000000')
                                        ws.cell(row=rowNum,column=2,value=str(PortKey)).font = Font(color='FF000000')
                                        ws.cell(row=rowNum,column=3,value="NA").font = Font(color='000000FF')
                                        if libPortDir.strip():
                                            ws.cell(row=rowNum,column=4,value=str(libPortDir.upper())).font = Font(color='FF000000')
                                        else:
                                            ws.cell(row=rowNum,column=4,value="ERROR").font = Font(color='00FF0000')
                                            # ERROR COMMENT
                                            if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                                ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                            else :
                                                ExistComment = ""
                                            ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {libFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in lib")).font = Font(color='FF000000')
                                            #                                             
                                        ws.cell(row=rowNum,column=6,value="FAIL").font = Font(color='00FF0000')
                                        # COMMENT
                                        if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                            ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                        else :
                                            ExistComment = ""
                                        ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {libFile}\nERROR : Cell '{CellKey}' defined in lib but not in lef")).font = Font(color='FF000000')                                            
                                        FoundExistCell = True
                                        break 
                                if not FoundExistCell:
                                    rowNum = len(ws['A'])+1
                                    ws.cell(row=rowNum,column=1,value=str(CellKey)).font = Font(color='FF000000')
                                    ws.cell(row=rowNum,column=2,value=str(PortKey)).font = Font(color='FF000000')
                                    ws.cell(row=rowNum,column=3,value="NA").font = Font(color='000000FF')
                                    if libPortDir.strip():
                                        ws.cell(row=rowNum,column=4,value=str(libPortDir.upper())).font = Font(color='FF000000')
                                    else:
                                        ws.cell(row=rowNum,column=4,value="ERROR").font = Font(color='00FF0000')
                                        # ERROR COMMENT
                                        if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                            ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                        else :
                                            ExistComment = ""
                                        ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {libFile}\nERROR : For Cell '{CellKey}' Pin '{PortKey}' direction not defined in lib")).font = Font(color='FF000000')
                                        #                                         
                                    ws.cell(row=rowNum,column=6,value="FAIL").font = Font(color='00FF0000')
                                    # COMMENT
                                    if ws.cell(rowNum,len(ExcelKeys)).value != None:
                                        ExistComment = ws.cell(rowNum,len(ExcelKeys)).value + "\n"
                                    else :
                                        ExistComment = ""
                                    ws.cell(row=rowNum,column=len(ExcelKeys),value=str(f"{ExistComment}File : {lefFile} Vs {libFile}\nERROR : Cell '{CellKey}' defined in lib but not in lef")).font = Font(color='FF000000')                                            
                            #    
    return wb,ws